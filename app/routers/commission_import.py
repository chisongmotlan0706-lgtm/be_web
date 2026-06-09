import logging
import math
import unicodedata
import uuid
from datetime import datetime, time, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.auth import get_current_user, user_owner_global_zalo
from app.commission_report import (
    parse_and_aggregate_report,
    parse_bill_conversion_completed_order_ids,
)
from app.db import get_supabase_client
from pydantic import BaseModel, Field

router = APIRouter(prefix="/commission-report", tags=["commission-report"])
logger = logging.getLogger(__name__)

MAX_UPLOAD_BYTES = 15 * 1024 * 1024
PAYOUT_IMPORT_MAX_BYTES = 5 * 1024 * 1024
_PAYOUT_FLOAT_VND_SLACK = 0.51
UPSERT_CHUNK = 500
LOOKUP_CHUNK = 500
PLACED_WITHIN_DAYS_ALLOWED = frozenset({1, 3, 7, 14})
ORDER_STATUS_COMPLETED_SYNC = "Hoàn thành"
ORDER_STATUS_PAID_SYNCED = "Đã cộng tiền"
_VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

SPLIT_TABLE = "affiliate_commission_order_splits"
SPLIT_ROLE_AGENCY = "agency"
SPLIT_ROLE_PLATFORM_OWNER = "platform_owner"
CONFIG_KEY_CHUYEN_TIEN = "chuyen_tien"
CONFIG_KEY_HOA_HONG = "hoa_hong"
_DEFAULT_NOI_DUNG_TEMPLATE = "Ho tro hoa hong {id_global}"
PAYOUT_STATUS_PAID = "Đã Trả Hoa Hồng"
WITHDRAW_STATUS_CHUA_BAO_KHACH = "CHUA_BAO_KHACH"
STATUS_BANK_LOI = "LOI_BANK"


def _placed_within_vn_calendar_to_now(days: int) -> tuple[datetime, datetime]:
    """
    Khoang [start, end] theo UTC:
    - start: 00:00 VN tai ngay (hom_nay_theo_lich_VN - days)
    - end: thoi diem hien tai UTC (gom ca phan hom nay da troi qua).
    """
    now_utc = datetime.now(timezone.utc)
    today_vn = now_utc.astimezone(_VN_TZ).date()
    start_date_vn = today_vn - timedelta(days=days)
    start_local = datetime.combine(start_date_vn, time.min, tzinfo=_VN_TZ)
    start_utc = start_local.astimezone(timezone.utc)
    return start_utc, now_utc


def _chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def _fetch_existing_order_status_by_order_id(
    supabase, order_ids: list[str]
) -> dict[str, str | None]:
    """order_id -> order_status trong DB truoc lan import nay."""
    out: dict[str, str | None] = {}
    unique_ids = sorted({oid.strip() for oid in order_ids if str(oid).strip()})
    for chunk in _chunked(unique_ids, LOOKUP_CHUNK):
        result = (
            supabase.table("affiliate_commission_orders")
            .select("order_id,order_status")
            .in_("order_id", chunk)
            .execute()
        )
        for item in result.data or []:
            oid = str(item.get("order_id") or "").strip()
            if not oid:
                continue
            out[oid] = str(item.get("order_status") or "").strip() or None
    return out


def _apply_order_status_transition(rows: list[dict], existing_status: dict[str, str | None]) -> None:
    """
    order_status luon lay tu file (da co trong row).
    order_status_transition: chi khi doi trang thai (cũ -> mới); khong doi thi null.
    """
    for row in rows:
        oid = str(row.get("order_id") or "").strip()
        new_st = str(row.get("order_status") or "").strip()
        old_st = existing_status.get(oid)
        if old_st is None:
            row["order_status_transition"] = None
            continue
        if old_st != new_st:
            row["order_status_transition"] = f"{old_st} -> {new_st}"
        else:
            row["order_status_transition"] = None


def _build_convert_info_map(supabase, sub_id_values: list[str]) -> dict[str, dict[str, str | None]]:
    """
    Map sub_id1 (= convert_results.id_zl) -> {id_globalzalo, id_globalgroup}.
    File Shopee sub_id1 trung id_zl; global lay tu cot da backfill tren convert_results.
    """
    mapping: dict[str, dict[str, str | None]] = {}

    def _ingest_convert_rows(items: list[dict]) -> None:
        for item in items or []:
            id_zl = str(item.get("id_zl") or "").strip()
            if not id_zl:
                continue
            id_gz = str(item.get("id_globalzalo") or "").strip() or None
            id_ggrp = str(item.get("id_globalgroup") or "").strip() or None
            mapping[id_zl] = {
                "id_globalzalo": id_gz,
                "id_globalgroup": id_ggrp,
            }

    for chunk in _chunked(sub_id_values, LOOKUP_CHUNK):
        result = (
            supabase.table("convert_results")
            .select("id_zl,id_globalzalo,id_globalgroup")
            .in_("id_zl", chunk)
            .execute()
        )
        _ingest_convert_rows(result.data or [])
    return mapping


def _build_contact_map_by_global(supabase, id_global_values: list[str]) -> dict[str, dict[str, str | None]]:
    """Map zalo_contacts.id_global -> {id_global, name}."""
    mapping: dict[str, dict[str, str | None]] = {}
    unique = sorted({str(x or "").strip() for x in id_global_values if str(x or "").strip()})
    for chunk in _chunked(unique, LOOKUP_CHUNK):
        result = (
            supabase.table("zalo_contacts")
            .select("id_global,d_name")
            .in_("id_global", chunk)
            .execute()
        )
        for item in result.data or []:
            id_g = str(item.get("id_global") or "").strip()
            if not id_g:
                continue
            name_raw = item.get("d_name")
            mapping[id_g] = {
                "id_global": id_g,
                "name": str(name_raw).strip() if name_raw is not None else None,
            }
    return mapping


def _resolve_contact_global(supabase, id_global_key: str | None) -> str | None:
    """Xac nhan id_global ton tai trong zalo_contacts."""
    k = str(id_global_key or "").strip()
    if not k:
        return None
    try:
        res = (
            supabase.table("zalo_contacts")
            .select("id_global")
            .eq("id_global", k)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if rows:
            v = str(rows[0].get("id_global") or "").strip()
            return v or None
    except Exception:
        logger.exception("[commission-import] resolve contact id_global failed key=%s", k)
    return None


def _fetch_group_globals_for_import(
    supabase, id_globalgroup: str
) -> tuple[str | None, str | None, str | None]:
    """
    Tra zalo_groups theo id_global (= id_globalgroup tu convert_results).
    Tra (group_id_global, agency_id_global, owner_id_global_main).

    Dai ly (split agency id_global):
      id_globalgroup -> zalo_groups.id_global -> zalo_groups.id_globalzalo.

    Chu tool (split platform_owner id_global): zalo_groups.id_global_main.
    """
    gid = str(id_globalgroup or "").strip()
    if not gid:
        return None, None, None
    res = (
        supabase.table("zalo_groups")
        .select("id_global,id_global_main,id_globalzalo")
        .eq("id_global", gid)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
    )
    rows = res.data or []
    if not rows:
        return None, None, None
    row = rows[0]
    group_id_global = str(row.get("id_global") or "").strip() or None
    owner_id_global_main = str(row.get("id_global_main") or "").strip() or None
    agency_id_global = str(row.get("id_globalzalo") or "").strip() or None
    return group_id_global, agency_id_global, owner_id_global_main


def _strip_legacy_import_order_fields(row: dict) -> dict:
    """Don import: chi id_global; xoa id_zl legacy tren ban ghi cu khi upsert."""
    out = dict(row)
    out.pop("id_zl", None)
    out["id_zl"] = None
    return out


def _split_row_for_import_upsert(row: dict) -> dict:
    """Split import: chi id_global; khong ghi id_zl (null de xoa legacy)."""
    out = dict(row)
    out["id_zl"] = None
    return out


def _clamp_pct_0_100(x: float) -> float:
    if not math.isfinite(x):
        return 0.0
    return max(0.0, min(100.0, x))


def _parse_app_config_pct_field(raw: object, *, default: float = 0.0) -> float:
    s = str(raw or "").strip().replace(",", "")
    if not s:
        return default
    try:
        return _clamp_pct_0_100(float(s))
    except (ValueError, TypeError):
        return default


def _fetch_hoa_hong_pct_quads(supabase) -> tuple[float, float, float, float, bool]:
    """
    app_config_kv.config_key = hoa_hong, is_active:
    - value_1: agency % (0-100)
    - value_2: owner (platform) % (0-100)
    - value_3: user hh % (0-100)
    - value_4: tru % (0-100) — he so (100 - value_4) / 100 tren net
    Tra (v1, v2, v3, v4, ok). ok = co dong active.
    """
    try:
        result = (
            supabase.table("app_config_kv")
            .select("value_1,value_2,value_3,value_4,is_active")
            .eq("config_key", CONFIG_KEY_HOA_HONG)
            .limit(1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            return (0.0, 0.0, 0.0, 0.0, False)
        row = rows[0]
        if row.get("is_active") is False:
            return (0.0, 0.0, 0.0, 0.0, False)
        v1 = _parse_app_config_pct_field(row.get("value_1"))
        v2 = _parse_app_config_pct_field(row.get("value_2"))
        v3 = _parse_app_config_pct_field(row.get("value_3"))
        v4 = _parse_app_config_pct_field(row.get("value_4"))
        return (v1, v2, v3, v4, True)
    except Exception:
        logger.exception("[commission-import] fetch hoa_hong app_config_kv failed")
        return (0.0, 0.0, 0.0, 0.0, False)


def _amounts_from_hoa_hong_net(
    net: float, v1: float, v2: float, v3: float, v4: float
) -> tuple[float, float, float]:
    """agency, owner, hh_user — cong thuc chuan: net * (100-v4) * vk / 10000."""
    if not math.isfinite(net):
        net = 0.0
    f = (100.0 - v4) / 10000.0
    return (
        round(net * f * v1, 4),
        round(net * f * v2, 4),
        round(net * f * v3, 4),
    )


def _effective_split_payout_pct_on_net(role_pct: float, v4: float) -> float:
    """% cua net tuong ung amount (de luu payout_pct tren bang splits)."""
    return round((100.0 - v4) * role_pct / 100.0, 4)


def _is_order_status_cancelled(order_status: str | None) -> bool:
    t = (order_status or "").strip().lower()
    return "hủy" in t or "huỷ" in t


def _fetch_group_owner_and_agency_ids(
    supabase, group_lookup_key: str
) -> tuple[str | None, str | None, str | None, str | None]:
    """
    Tra zalo_groups theo id_global (uu tien) roi group_id (local).
    Tra (agency_id_zl, owner_id_zl_main, agency_id_global, owner_id_global_main).
    """
    gid = str(group_lookup_key or "").strip()
    if not gid:
        return None, None, None, None
    row = None
    for col in ("id_global", "group_id"):
        res = (
            supabase.table("zalo_groups")
            .select("id_zl,id_zl_main,id_global,id_global_main")
            .eq(col, gid)
            .is_("deleted_at", "null")
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if rows:
            row = rows[0]
            break
    if not row:
        return None, None, None, None
    return (
        str(row.get("id_zl") or "").strip() or None,
        str(row.get("id_zl_main") or "").strip() or None,
        str(row.get("id_global") or "").strip() or None,
        str(row.get("id_global_main") or "").strip() or None,
    )


def _fetch_commission_orders_rows_by_order_ids(
    supabase, order_ids: list[str]
) -> list[dict]:
    out: list[dict] = []
    unique_ids = sorted({oid.strip() for oid in order_ids if str(oid).strip()})
    for chunk in _chunked(unique_ids, LOOKUP_CHUNK):
        result = (
            supabase.table("affiliate_commission_orders")
            .select("id,order_id,order_status,net_affiliate_commission,sub_id1,source_filename")
            .in_("order_id", chunk)
            .execute()
        )
        for item in result.data or []:
            out.append(item)
    return out


def _fetch_orders_map_by_order_id(supabase, order_ids: list[str]) -> dict[str, dict]:
    """order_id -> ban ghi day du (select *) de so sanh truoc khi upsert."""
    out: dict[str, dict] = {}
    unique_ids = sorted({oid.strip() for oid in order_ids if str(oid).strip()})
    for chunk in _chunked(unique_ids, LOOKUP_CHUNK):
        result = (
            supabase.table("affiliate_commission_orders")
            .select("*")
            .in_("order_id", chunk)
            .execute()
        )
        for item in result.data or []:
            oid = str(item.get("order_id") or "").strip()
            if oid:
                out[oid] = item
    return out


def _fetch_splits_detail_by_commission_order_ids(
    supabase, commission_ids: list[str]
) -> dict[str, dict[str, dict[str, str | float] | None]]:
    """
    commission_order_id -> { agency: {id_global, amount} | None, owner: {id_global, amount} | None }
    """
    out: dict[str, dict[str, dict[str, str | float] | None]] = {}
    unique = sorted({str(x).strip() for x in commission_ids if str(x).strip()})
    for chunk in _chunked(unique, LOOKUP_CHUNK):
        result = (
            supabase.table(SPLIT_TABLE)
            .select("commission_order_id,split_role,id_global,amount")
            .in_("commission_order_id", chunk)
            .execute()
        )
        for item in result.data or []:
            cid = str(item.get("commission_order_id") or "").strip()
            if not cid:
                continue
            if cid not in out:
                out[cid] = {"agency": None, "owner": None}
            role = str(item.get("split_role") or "").strip()
            entry: dict[str, str | float] = {
                "id_global": str(item.get("id_global") or "").strip(),
                "amount": round(float(item.get("amount") or 0), 4),
            }
            if role == SPLIT_ROLE_AGENCY:
                out[cid]["agency"] = entry
            elif role == SPLIT_ROLE_PLATFORM_OWNER:
                out[cid]["owner"] = entry
    return out


def _contact_display_name(contact_map: dict[str, dict], id_key: str | None) -> str | None:
    if not id_key:
        return None
    info = contact_map.get(str(id_key).strip())
    if not info:
        return None
    name = info.get("name")
    if name is not None and str(name).strip():
        return str(name).strip()
    return None


def _build_bill_sync_hh_preview(supabase, order_ids: list[str]) -> list[dict]:
    """order_ids: thu tu hien thi nhu file (unique)."""
    if not order_ids:
        return []
    db_map = _fetch_orders_map_by_order_id(supabase, order_ids)
    commission_ids: list[str] = []
    id_globals: list[str] = []
    for oid in order_ids:
        row = db_map.get(oid)
        if not row:
            continue
        raw_id = row.get("id")
        if raw_id is not None:
            commission_ids.append(str(raw_id))
        id_global = str(row.get("id_global") or "").strip()
        if id_global:
            id_globals.append(id_global)
    split_detail = _fetch_splits_detail_by_commission_order_ids(supabase, commission_ids)
    for detail in split_detail.values():
        for role_key in ("agency", "owner"):
            part = detail.get(role_key)
            if part and part.get("id_global"):
                id_globals.append(str(part["id_global"]).strip())
    contact_map = _build_contact_map_by_global(supabase, sorted(set(id_globals)))

    empty_payout = {
        "user_name": None,
        "user_amount": 0.0,
        "agency_name": None,
        "agency_amount": 0.0,
        "owner_name": None,
        "owner_amount": 0.0,
    }

    out: list[dict] = []
    for oid in order_ids:
        row = db_map.get(oid)
        if not row:
            out.append(
                {
                    "order_id": oid,
                    "db_found": False,
                    "db_order_status": None,
                    "id_global": None,
                    "has_zalo_contact": False,
                    "eligible": False,
                    "skip_reason": "Khong co trong DB (chua import)",
                    **empty_payout,
                }
            )
            continue

        db_st = str(row.get("order_status") or "").strip()
        id_global = str(row.get("id_global") or "").strip() or None
        has_contact = bool(id_global and id_global in contact_map)
        cid = str(row.get("id") or "").strip()
        hh_user = round(float(row.get("hh_user") or 0), 4)

        skip_reason: str | None = None
        eligible = False
        if db_st == ORDER_STATUS_PAID_SYNCED:
            skip_reason = "Da cong tien trong DB"
        elif db_st != ORDER_STATUS_COMPLETED_SYNC:
            skip_reason = f"Trang thai DB khac Hoan thanh ({db_st or 'rong'})"
        elif not id_global:
            skip_reason = "Thieu id_global (affiliate)"
        elif not has_contact:
            skip_reason = "Khong co zalo_contacts theo id_global"
        else:
            eligible = True

        user_name = _contact_display_name(contact_map, id_global)
        if not user_name and row.get("name"):
            user_name = str(row.get("name") or "").strip() or None
        user_amount = hh_user if eligible else 0.0

        agency_name: str | None = None
        agency_amount = 0.0
        owner_name: str | None = None
        owner_amount = 0.0
        if cid:
            parts = split_detail.get(cid) or {}
            agency_part = parts.get("agency")
            if agency_part:
                agency_global = str(agency_part.get("id_global") or "").strip()
                agency_name = _contact_display_name(contact_map, agency_global)
                if agency_global and agency_global in contact_map and eligible:
                    agency_amount = float(agency_part.get("amount") or 0)
            owner_part = parts.get("owner")
            if owner_part:
                owner_global = str(owner_part.get("id_global") or "").strip()
                owner_name = _contact_display_name(contact_map, owner_global)
                if owner_global and owner_global in contact_map and eligible:
                    owner_amount = float(owner_part.get("amount") or 0)

        out.append(
            {
                "order_id": oid,
                "db_found": True,
                "db_order_status": db_st,
                "id_global": id_global,
                "has_zalo_contact": has_contact,
                "eligible": eligible,
                "skip_reason": skip_reason,
                "user_name": user_name,
                "user_amount": round(user_amount, 4),
                "agency_name": agency_name,
                "agency_amount": round(agency_amount, 4),
                "owner_name": owner_name,
                "owner_amount": round(owner_amount, 4),
            }
        )
    return out


def _fetch_orders_for_summary(
    supabase,
    *,
    placed_within_days: int | None,
    page_size: int = 1000,
) -> list[dict]:
    out: list[dict] = []
    start_utc = None
    end_utc = None
    if placed_within_days is not None:
        start_utc, end_utc = _placed_within_vn_calendar_to_now(placed_within_days)
    offset = 0
    while True:
        query = (
            supabase.table("affiliate_commission_orders")
            .select("hh_user,order_status")
            .order("id")
            .range(offset, offset + page_size - 1)
        )
        if start_utc is not None and end_utc is not None:
            query = query.gte("order_placed_at", start_utc.isoformat()).lte(
                "order_placed_at", end_utc.isoformat()
            )
        result = query.execute()
        rows = result.data or []
        if not rows:
            break
        out.extend(rows)
        if len(rows) < page_size:
            break
        offset += page_size
    return out


def _fetch_splits_for_summary(
    supabase,
    *,
    placed_within_days: int | None,
    page_size: int = 1000,
) -> list[dict]:
    out: list[dict] = []
    start_utc = None
    end_utc = None
    if placed_within_days is not None:
        start_utc, end_utc = _placed_within_vn_calendar_to_now(placed_within_days)
    offset = 0
    while True:
        query = (
            supabase.table(SPLIT_TABLE)
            .select("split_role,amount,order_status")
            .order("id")
            .range(offset, offset + page_size - 1)
        )
        if start_utc is not None and end_utc is not None:
            query = query.gte("created_at", start_utc.isoformat()).lte(
                "created_at", end_utc.isoformat()
            )
        result = query.execute()
        rows = result.data or []
        if not rows:
            break
        out.extend(rows)
        if len(rows) < page_size:
            break
        offset += page_size
    return out


def _fetch_zalo_contacts_available_total(supabase) -> float:
    total = 0.0
    offset = 0
    page_size = 1000
    while True:
        result = (
            supabase.table("zalo_contacts")
            .select("available_amount")
            .order("id_global")
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            break
        for row in rows:
            total += float(row.get("available_amount") or 0)
        if len(rows) < page_size:
            break
        offset += page_size
    return total


def _is_admin_zalo_contact_role(role: object) -> bool:
    """ADMIN = chu tool / quan tri; khong hien thi trong danh sach user + dai ly."""
    return str(role or "").strip().upper() == "ADMIN"


def _fetch_chuyen_tien_settings(supabase) -> tuple[float, str]:
    """
    app_config_kv.config_key = chuyen_tien:
    - value_1: nguong VND toi thieu (available_amount) de hien thi / export.
    - value_2: tuy chon — mau noi dung CK; ho tro {id_global}, {d_name}. Neu trong -> mac dinh.
    Chi ap nguong khi dong ton tai, is_active, value_1 parse duoc > 0.
    """
    try:
        result = (
            supabase.table("app_config_kv")
            .select("value_1,value_2,is_active")
            .eq("config_key", CONFIG_KEY_CHUYEN_TIEN)
            .limit(1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            return (0.0, _DEFAULT_NOI_DUNG_TEMPLATE)
        row = rows[0]
        if row.get("is_active") is False:
            return (0.0, _DEFAULT_NOI_DUNG_TEMPLATE)
        raw = str(row.get("value_1") or "").strip().replace(",", "")
        min_vnd = max(0.0, float(raw)) if raw else 0.0
        v2 = str(row.get("value_2") or "").strip()
        template = v2 if v2 else _DEFAULT_NOI_DUNG_TEMPLATE
        return (min_vnd, template)
    except (ValueError, TypeError, Exception):
        return (0.0, _DEFAULT_NOI_DUNG_TEMPLATE)


def _format_noi_dung_ck(template: str, *, id_global: str, d_name: str) -> str:
    safe_name = (d_name or "").strip() or id_global
    try:
        return str(template).format(id_global=id_global, d_name=safe_name)
    except (KeyError, ValueError, IndexError):
        return (
            str(template)
            .replace("{id_global}", id_global)
            .replace("{d_name}", safe_name)
        )


def _fetch_group_ids_by_owner(
    supabase, *, owner_id_global_main: str, active_only: bool
) -> list[str]:
    """id_global tu zalo_groups: id_global_main = owner (auth id_globalzalo), chua xoa mem."""
    out: list[str] = []
    seen: set[str] = set()
    owner = str(owner_id_global_main or "").strip()
    if not owner:
        return out
    offset = 0
    page_size = 1000
    while True:
        query = (
            supabase.table("zalo_groups")
            .select("id_global")
            .eq("id_global_main", owner)
            .is_("deleted_at", "null")
            .order("id")
            .range(offset, offset + page_size - 1)
        )
        if active_only:
            query = query.eq("status", "ACTIVE")
        result = query.execute()
        rows = result.data or []
        if not rows:
            break
        for row in rows:
            gid = str(row.get("id_global") or "").strip()
            if gid and gid not in seen:
                seen.add(gid)
                out.append(gid)
        if len(rows) < page_size:
            break
        offset += page_size
    return out


def _fetch_contact_globals_for_group_ids(supabase, group_global_ids: list[str]) -> set[str]:
    """id_global cua zalo_contacts co id_global_gr thuoc danh sach id_global nhom."""
    out: set[str] = set()
    unique_groups = sorted(
        {str(g or "").strip() for g in group_global_ids if str(g or "").strip()}
    )
    if not unique_groups:
        return out
    page_size = 1000
    for chunk in _chunked(unique_groups, LOOKUP_CHUNK):
        offset = 0
        while True:
            result = (
                supabase.table("zalo_contacts")
                .select("id_global")
                .in_("id_global_gr", chunk)
                .order("id")
                .range(offset, offset + page_size - 1)
                .execute()
            )
            rows = result.data or []
            if not rows:
                break
            for row in rows:
                zid = str(row.get("id_global") or "").strip()
                if zid:
                    out.add(zid)
            if len(rows) < page_size:
                break
            offset += page_size
    return out


def _fetch_zalo_contacts_by_id_groups(
    supabase,
    *,
    group_global_ids: list[str],
    limit: int,
    min_available_amount: float | None = None,
) -> list[dict]:
    """limit > 0: cat sau sort updated_at. limit = 0: lay het (de sort khac o buoc enrich)."""
    if limit < 0:
        return []
    unique_groups = sorted(
        {str(g or "").strip() for g in group_global_ids if str(g or "").strip()}
    )
    if not unique_groups:
        return []
    allowed = set(unique_groups)
    out: list[dict] = []
    for chunk in _chunked(unique_groups, LOOKUP_CHUNK):
        query = (
            supabase.table("zalo_contacts")
            .select(
                "id,id_global,d_name,id_global_gr,available_amount,actual_amount,estimated_amount,"
                "role,bank_name,bank_type,stk,status_bank,updated_at,received"
            )
            .in_("id_global_gr", chunk)
        )
        if min_available_amount is not None and min_available_amount > 0:
            query = query.gte("available_amount", min_available_amount)
        result = query.execute()
        for item in result.data or []:
            gid = str(item.get("id_global_gr") or "").strip()
            if gid in allowed:
                out.append(item)
    out.sort(
        key=lambda item: str(item.get("updated_at") or ""),
        reverse=True,
    )
    if limit > 0:
        return out[:limit]
    return out


def _fetch_order_balance_maps_by_id_zl(
    supabase, *, allowed_contact_global_ids: set[str] | None = None
) -> tuple[dict[str, float], dict[str, float]]:
    pending_map: dict[str, float] = {}
    completed_map: dict[str, float] = {}
    allowed = allowed_contact_global_ids or set()
    offset = 0
    page_size = 1000
    while True:
        result = (
            supabase.table("affiliate_commission_orders")
            .select("id_global,hh_user,order_status")
            .order("id")
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            break
        for row in rows:
            id_g = str(row.get("id_global") or "").strip()
            if not id_g:
                continue
            if allowed and id_g not in allowed:
                continue
            status = str(row.get("order_status") or "").strip()
            if _is_order_status_cancelled(status):
                continue
            amount = float(row.get("hh_user") or 0)
            if status == "Đang chờ xử lý":
                pending_map[id_g] = pending_map.get(id_g, 0.0) + amount
            elif status == "Hoàn thành":
                completed_map[id_g] = completed_map.get(id_g, 0.0) + amount
        if len(rows) < page_size:
            break
        offset += page_size
    return pending_map, completed_map


def _fetch_split_balance_maps_by_id_zl(
    supabase, active_member_global_ids: set[str]
) -> tuple[dict[str, float], dict[str, float]]:
    pending_map: dict[str, float] = {}
    completed_map: dict[str, float] = {}
    offset = 0
    page_size = 1000
    while True:
        result = (
            supabase.table(SPLIT_TABLE)
            .select("id_global,amount,order_status")
            .order("id")
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            break
        for row in rows:
            id_g = str(row.get("id_global") or "").strip()
            if not id_g or id_g not in active_member_global_ids:
                continue
            status = str(row.get("order_status") or "").strip()
            if _is_order_status_cancelled(status):
                continue
            amount = float(row.get("amount") or 0)
            if status == "Đang chờ xử lý":
                pending_map[id_g] = pending_map.get(id_g, 0.0) + amount
            elif status == "Hoàn thành":
                completed_map[id_g] = completed_map.get(id_g, 0.0) + amount
        if len(rows) < page_size:
            break
        offset += page_size
    return pending_map, completed_map


def _incoming_commission_row_equals_db(incoming: dict, db: dict) -> bool:
    """
    True neu order_status tu file trung voi DB -> khong co "thay doi trang thai" -> bo qua upsert.
    Net/thoi gian/sub_id1/hh_user/... khong tinh la thay doi (chi can status khac moi upsert).
    Chi goi khi da co ban ghi DB (don moi khong goi).
    """
    return str(incoming.get("order_status") or "").strip() == str(db.get("order_status") or "").strip()


def _partition_rows_skip_unchanged(supabase, rows: list[dict]) -> tuple[list[dict], int]:
    """
    Bo qua upsert neu order_status file == DB (khong doi trang thai).
    Tra ve (rows_se_upsert, so_dong_bo_qua).
    """
    if not rows:
        return [], 0
    order_ids = [
        str(r.get("order_id") or "").strip() for r in rows if str(r.get("order_id") or "").strip()
    ]
    existing = _fetch_orders_map_by_order_id(supabase, order_ids)
    out: list[dict] = []
    skipped = 0
    for r in rows:
        oid = str(r.get("order_id") or "").strip()
        db_row = existing.get(oid)
        if db_row is not None and _incoming_commission_row_equals_db(r, db_row):
            skipped += 1
            continue
        out.append(r)
    return out, skipped


def _sync_commission_order_splits_for_import(
    supabase,
    orders_from_db: list[dict],
    import_batch_id: str,
    convert_info_map: dict[str, dict[str, str | None]],
    *,
    hoa_v1: float,
    hoa_v2: float,
    hoa_v4: float,
) -> dict[str, int]:
    """
    Dong bo affiliate_commission_order_splits sau khi upsert don chinh.
    Don Da huy: chi cap nhat order_status tren splits, giu amount.
    Don Da cong tien: khong goi ham nay.
    """
    stats = {
        "splits_rows_upserted": 0,
        "orders_splits_status_only": 0,
        "splits_rows_deleted": 0,
    }
    for order in orders_from_db:
        cid_raw = order.get("id")
        oid = str(order.get("order_id") or "").strip()
        if not cid_raw or not oid:
            continue
        cid = str(cid_raw)
        order_status = str(order.get("order_status") or "").strip()
        net = float(order.get("net_affiliate_commission") or 0)
        sub_id1 = str(order.get("sub_id1") or "").strip() or None
        source_filename = order.get("source_filename")

        if _is_order_status_cancelled(order_status):
            try:
                supabase.table(SPLIT_TABLE).update(
                    {
                        "order_status": order_status,
                        "import_batch_id": import_batch_id,
                        "amount": 0,
                        "net_affiliate_commission_at_split": 0,
                    }
                ).eq("commission_order_id", cid).execute()
            except Exception as exc:
                logger.exception(
                    "[commission-import] splits cancel status update failed order_id=%s",
                    oid,
                )
                raise HTTPException(
                    status_code=500,
                    detail=f"Cap nhat splits (Da huy) loi: {exc}",
                ) from exc
            stats["orders_splits_status_only"] += 1
            continue

        if not sub_id1:
            try:
                del_res = (
                    supabase.table(SPLIT_TABLE).delete().eq("commission_order_id", cid).execute()
                )
                n = len(del_res.data or []) if del_res.data is not None else 0
                stats["splits_rows_deleted"] += n
            except Exception as exc:
                logger.exception(
                    "[commission-import] splits delete (missing sub_id1) failed order_id=%s",
                    oid,
                )
                raise HTTPException(
                    status_code=500,
                    detail=f"Xoa splits loi: {exc}",
                ) from exc
            continue

        convert_info = convert_info_map.get(sub_id1)
        id_globalgroup = (
            str(convert_info.get("id_globalgroup") or "").strip() or None if convert_info else None
        )
        group_id_global: str | None = None
        agency_id_global: str | None = None
        owner_id_global_main: str | None = None
        if id_globalgroup:
            group_id_global, agency_id_global, owner_id_global_main = _fetch_group_globals_for_import(
                supabase, id_globalgroup
            )

        if not id_globalgroup:
            raise HTTPException(
                status_code=500,
                detail=(
                    f"Khong tim thay id_globalgroup (convert_results) cho order_id={oid} "
                    f"(sub_id1={sub_id1})"
                ),
            )
        if not owner_id_global_main:
            raise HTTPException(
                status_code=500,
                detail=(
                    f"Khong tim thay id_global_main trong zalo_groups cho id_globalgroup="
                    f"{id_globalgroup} (order_id={oid})"
                ),
            )

        agency_amount, owner_amount, _ = _amounts_from_hoa_hong_net(
            net, hoa_v1, hoa_v2, 0.0, hoa_v4
        )
        agency_pct_stored = _effective_split_payout_pct_on_net(hoa_v1, hoa_v4)
        owner_pct_stored = _effective_split_payout_pct_on_net(hoa_v2, hoa_v4)

        split_group_id = group_id_global or id_globalgroup

        split_rows: list[dict] = []
        if agency_id_global:
            split_rows.append(
                {
                    "commission_order_id": cid,
                    "order_id": oid,
                    "split_role": SPLIT_ROLE_AGENCY,
                    "id_global": agency_id_global,
                    "group_id": split_group_id,
                    "payout_pct": agency_pct_stored,
                    "amount": agency_amount,
                    "net_affiliate_commission_at_split": net,
                    "order_status": order_status,
                    "import_batch_id": import_batch_id,
                    "source_filename": source_filename,
                }
            )
        split_rows.append(
            {
                "commission_order_id": cid,
                "order_id": oid,
                "split_role": SPLIT_ROLE_PLATFORM_OWNER,
                "id_global": owner_id_global_main,
                "group_id": split_group_id,
                "payout_pct": owner_pct_stored,
                "amount": owner_amount,
                "net_affiliate_commission_at_split": net,
                "order_status": order_status,
                "import_batch_id": import_batch_id,
                "source_filename": source_filename,
            }
        )

        try:
            supabase.table(SPLIT_TABLE).upsert(
                [_split_row_for_import_upsert(r) for r in split_rows],
                on_conflict="commission_order_id,split_role",
            ).execute()
            stats["splits_rows_upserted"] += len(split_rows)
        except Exception as exc:
            logger.exception("[commission-import] splits upsert failed order_id=%s", oid)
            raise HTTPException(
                status_code=500,
                detail=f"Luu splits loi: {exc}",
            ) from exc

        agency_row_present = any(
            r.get("split_role") == SPLIT_ROLE_AGENCY for r in split_rows
        )
        if not agency_row_present:
            try:
                del_res = (
                    supabase.table(SPLIT_TABLE)
                    .delete()
                    .eq("commission_order_id", cid)
                    .eq("split_role", SPLIT_ROLE_AGENCY)
                    .execute()
                )
                n = len(del_res.data or []) if del_res.data is not None else 0
                stats["splits_rows_deleted"] += n
            except Exception as exc:
                logger.exception(
                    "[commission-import] splits delete orphan agency failed order_id=%s",
                    oid,
                )
                raise HTTPException(
                    status_code=500,
                    detail=f"Xoa split agency loi: {exc}",
                ) from exc

    return stats


@router.get("/summary")
def get_commission_summary(
    placed_within_days: int | None = Query(
        default=None,
        description=(
            "Loc du lieu theo khoang 1,3,7,14 ngay. Orders dung order_placed_at; splits dung created_at. "
            "Bo qua = tat ca."
        ),
    ),
):
    if placed_within_days is not None and placed_within_days not in PLACED_WITHIN_DAYS_ALLOWED:
        raise HTTPException(
            status_code=400,
            detail="placed_within_days chi chap nhan 1, 3, 7 hoac 14",
        )
    try:
        supabase = get_supabase_client()
        orders = _fetch_orders_for_summary(supabase, placed_within_days=placed_within_days)
        splits = _fetch_splits_for_summary(supabase, placed_within_days=placed_within_days)
        user_available = _fetch_zalo_contacts_available_total(supabase)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query summary loi: {exc}",
        ) from exc

    user_processing = 0.0
    user_completed = 0.0
    for row in orders:
        status = str(row.get("order_status") or "").strip()
        hh_user = float(row.get("hh_user") or 0)
        if _is_order_status_cancelled(status):
            continue
        if status == "Hoàn thành":
            user_completed += hh_user
        elif status != ORDER_STATUS_PAID_SYNCED:
            user_processing += hh_user

    agency_total = 0.0
    owner_total = 0.0
    for row in splits:
        status = str(row.get("order_status") or "").strip()
        if _is_order_status_cancelled(status):
            continue
        amount = float(row.get("amount") or 0)
        role = str(row.get("split_role") or "").strip()
        if role == SPLIT_ROLE_AGENCY:
            agency_total += amount
        elif role == SPLIT_ROLE_PLATFORM_OWNER:
            owner_total += amount

    return {
        "user": {
            "dang_xu_ly": round(user_processing, 4),
            "hoan_thanh": round(user_completed, 4),
            "co_san": round(user_available, 4),
        },
        "agency": {"tong": round(agency_total, 4)},
        "owner": {"tong": round(owner_total, 4)},
        "currency": "VND",
        "filters": {"placed_within_days": placed_within_days},
    }


def _enriched_zalo_contacts_for_owner(
    supabase,
    *,
    user_owner_global: str,
    limit: int,
    min_available_amount: float | None,
    sort_by: str = "updated_at",
    include_admin: bool = False,
) -> list[dict]:
    owner = str(user_owner_global or "").strip()
    if not owner:
        return []

    all_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=False
    )
    if not all_group_globals:
        return []

    allowed_contact_global = _fetch_contact_globals_for_group_ids(supabase, all_group_globals)
    if not allowed_contact_global:
        return []

    order_pending_map, order_completed_map = _fetch_order_balance_maps_by_id_zl(
        supabase, allowed_contact_global_ids=allowed_contact_global
    )
    active_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=True
    )
    active_member_global = (
        _fetch_contact_globals_for_group_ids(supabase, active_group_globals)
        if active_group_globals
        else set()
    )
    split_pending_map, split_completed_map = _fetch_split_balance_maps_by_id_zl(
        supabase, active_member_global
    )
    sort_key = str(sort_by or "updated_at").strip().lower()
    fetch_cap = 0 if sort_key == "tien_co_the_rut" else limit
    items = _fetch_zalo_contacts_by_id_groups(
        supabase,
        group_global_ids=all_group_globals,
        limit=fetch_cap,
        min_available_amount=min_available_amount,
    )
    if not include_admin:
        items = [it for it in items if not _is_admin_zalo_contact_role(it.get("role"))]
    for item in items:
        id_c = str(item.get("id_global") or "").strip()
        order_pending = order_pending_map.get(id_c, 0.0)
        order_completed = order_completed_map.get(id_c, 0.0)
        split_pending = split_pending_map.get(id_c, 0.0)
        split_completed = split_completed_map.get(id_c, 0.0)
        item["dang_giao_hang"] = round(order_pending + split_pending, 4)
        item["cho_duyet"] = round(order_completed + split_completed, 4)
        item["tien_co_the_rut"] = float(_floor_vnd_to_thousand(float(item.get("available_amount") or 0)))
        item["da_rut_ve_bank"] = round(float(item.get("received") or 0), 4)
    if sort_key == "tien_co_the_rut":
        items.sort(key=lambda x: float(x.get("tien_co_the_rut") or 0), reverse=True)
    if limit > 0:
        items = items[:limit]
    return items


def _zalo_contacts_transfer_workbook(items: list[dict], *, noi_dung_template: str) -> BytesIO:
    """Export CK: Ten nguoi huong = bank_name; cot bank_type = zalo_contacts.bank_type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chuyen tien"

    headers = [
        "STT",
        "id_global",
        "id_global_gr",
        "Số tiền chuyển",
        "TK hưởng",
        "Tên người hưởng",
        "bank_type",
        "Nội dung",
        "Hoàn Tiền chưa",
    ]
    ws.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(items, start=1):
        id_g = str(item.get("id_global") or "").strip()
        id_gr = str(item.get("id_global_gr") or "").strip()
        bank_name = str(item.get("bank_name") or "").strip()
        bank = str(item.get("bank_type") or "").strip()
        stk_raw = str(item.get("stk") or "").strip()
        amt = _floor_vnd_to_thousand(float(item.get("available_amount") or 0))
        # Noi dung chuyen khoan yeu cau co dinh, khong dau: "Hoan tien hoa hong <so_tien>".
        noi = f"Hoan tien hoa hong {amt}"
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=idx)
        c_id = ws.cell(row=row_num, column=2, value=id_g if id_g else "")
        c_id.number_format = "@"
        c_gid = ws.cell(row=row_num, column=3, value=id_gr if id_gr else "")
        c_gid.number_format = "@"
        ws.cell(row=row_num, column=4, value=amt)
        c_stk = ws.cell(row=row_num, column=5, value=stk_raw if stk_raw else "")
        c_stk.number_format = "@"
        ws.cell(row=row_num, column=6, value=bank_name if bank_name else id_g)
        ws.cell(row=row_num, column=7, value=bank)
        ws.cell(row=row_num, column=8, value=noi)
        ws.cell(row=row_num, column=9, value="")

    widths = (6, 18, 18, 16, 18, 28, 22, 40, 22)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@router.get("/zalo-contacts")
def list_zalo_contacts(
    limit: int = Query(default=500, ge=1, le=2000),
    apply_min_filter: bool = Query(
        default=True,
        description="True: chi contact co available_amount >= chuyen_tien.value_1 (neu co cau hinh).",
    ),
    include_admin: bool = Query(
        default=False,
        description="True: gom ca contact role ADMIN (mac dinh an ADMIN).",
    ),
    sort_by: str = Query(
        default="updated_at",
        description="updated_at | tien_co_the_rut",
    ),
    current_user: dict = Depends(get_current_user),
):
    sort_norm = str(sort_by or "updated_at").strip().lower()
    if sort_norm not in ("updated_at", "tien_co_the_rut"):
        raise HTTPException(
            status_code=400,
            detail="sort_by chi chap nhan updated_at hoac tien_co_the_rut",
        )
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        min_vnd, _tpl = _fetch_chuyen_tien_settings(supabase)
        min_arg = None
        if apply_min_filter and min_vnd > 0:
            min_arg = min_vnd
        items = _enriched_zalo_contacts_for_owner(
            supabase,
            user_owner_global=user_owner,
            limit=limit,
            min_available_amount=min_arg,
            sort_by=sort_norm,
            include_admin=include_admin,
        )
        filters: dict[str, float | str | bool | None] = {
            "apply_min_filter": apply_min_filter,
            "include_admin": include_admin,
            "sort_by": sort_norm,
        }
        if min_vnd > 0:
            filters["min_available_vnd_chuyen_tien"] = round(min_vnd, 4)
        else:
            filters["min_available_vnd_chuyen_tien"] = None
        return {"count": len(items), "items": items, "filters": filters}
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query zalo_contacts loi: {exc}",
        ) from exc


def _contact_row_owned_by_user(
    supabase,
    *,
    contact_id: int,
    user_owner_global: str,
) -> dict:
    owner = str(user_owner_global or "").strip()
    if not owner:
        raise HTTPException(
            status_code=400,
            detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
        )
    contacts = _all_zalo_contact_rows_for_owner(supabase, user_owner_global=owner)
    _, by_id = _contact_maps_for_payout(contacts)
    row = by_id.get(contact_id)
    if not row:
        raise HTTPException(
            status_code=404,
            detail=f"contact_id={contact_id} khong ton tai hoac khong thuoc ban",
        )
    return row


@router.patch("/zalo-contacts/{contact_id}/status-bank")
def mark_zalo_contact_loi_bank(
    contact_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Danh dau contact LOI_BANK (status_bank = LOI_BANK), chi contact thuoc nhom cua user."""
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        _contact_row_owned_by_user(
            supabase, contact_id=contact_id, user_owner_global=user_owner
        )
        res = (
            supabase.table("zalo_contacts")
            .update({"status_bank": STATUS_BANK_LOI})
            .eq("id", contact_id)
            .execute()
        )
        updated = res.data or []
        if not updated:
            raise HTTPException(
                status_code=409,
                detail=f"Khong cap nhat duoc contact_id={contact_id}",
            )
        row = updated[0]
        return {
            "contact_id": contact_id,
            "id_global": str(row.get("id_global") or "").strip() or None,
            "status_bank": str(row.get("status_bank") or STATUS_BANK_LOI),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Cap nhat status_bank loi: {exc}",
        ) from exc


@router.get("/zalo-contacts/export")
def export_zalo_contacts_xlsx(
    limit: int = Query(default=2000, ge=1, le=5000),
    include_admin: bool = Query(
        default=False,
        description="True: gom ca contact role ADMIN trong file xuat.",
    ),
    current_user: dict = Depends(get_current_user),
):
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        min_vnd, noi_tpl = _fetch_chuyen_tien_settings(supabase)
        min_arg = min_vnd if min_vnd > 0 else None
        items = _enriched_zalo_contacts_for_owner(
            supabase,
            user_owner_global=user_owner,
            limit=limit,
            min_available_amount=min_arg,
            include_admin=include_admin,
        )
        bio = _zalo_contacts_transfer_workbook(items, noi_dung_template=noi_tpl)
        data = bio.getvalue()
        filename_ascii = "Danh_sach_KH.xlsx"
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_ascii}"',
            },
        )
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Export zalo_contacts loi: {exc}",
        ) from exc


def _fold_col_name(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or "").strip().lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").replace(" ", "")


def _pick_id_global_column(columns: list[str]) -> str | None:
    for c in columns:
        f = _fold_col_name(c)
        if "id_global" in f or f == "idglobal":
            return c
    return None


def _pick_amount_column(columns: list[str]) -> str | None:
    for c in columns:
        f = _fold_col_name(c)
        # "Số tiền chuyển" -> fold "sotienchuyen" (tien, khong phai "sotientchuyen")
        if "sotienchuyen" in f or "sotientchuyen" in f:
            return c
    return None


def _pick_status_column(columns: list[str]) -> str | None:
    for c in columns:
        f = _fold_col_name(c)
        if "hoantienchua" in f or f == "trangthai" or "trangthai" in f:
            return c
    return None


def _payout_status_cell_norm(v: object) -> str:
    return " ".join(str(v if v is not None else "").strip().lower().split())


def _is_payout_paid_cell(v: object) -> bool:
    return _payout_status_cell_norm(v) == _payout_status_cell_norm(PAYOUT_STATUS_PAID)


def _all_zalo_contact_rows_for_owner(
    supabase, *, user_owner_global: str, max_rows: int = 20000
) -> list[dict]:
    owner = str(user_owner_global or "").strip()
    if not owner:
        return []
    all_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=False
    )
    if not all_group_globals:
        return []
    return _fetch_zalo_contacts_by_id_groups(
        supabase,
        group_global_ids=all_group_globals,
        limit=max_rows,
        min_available_amount=None,
    )


def _payout_deduct_vnd_for_contact(*, avail: float, amount_vnd: int) -> float | None:
    """So tien tru available va cong received — theo cot So tien chuyen trong file Excel."""
    return _resolve_payout_deduct_vnd(avail, amount_vnd)


def _payout_withdraw_eligible(row: dict) -> tuple[bool, str | None]:
    if not _withdraw_bank_fields_from_contact(row):
        return False, "thieu bank_type, stk hoac bank_name"
    return True, None


def _withdraw_bank_fields_from_contact(row: dict) -> tuple[str, str, str] | None:
    bank_type = str(row.get("bank_type") or "").strip()
    stk = str(row.get("stk") or "").strip()
    bank_name = str(row.get("bank_name") or "").strip()
    if not bank_type or not stk or not bank_name:
        return None
    return bank_type, stk, bank_name


def _insert_withdraw_request_chua_bao_khach(
    supabase,
    *,
    id_global: str,
    d_name: str | None,
    amount_vnd: int,
    bank_type: str,
    stk: str,
    bank_name: str,
) -> str:
    if amount_vnd <= 0:
        raise HTTPException(status_code=400, detail="amount withdraw_requests phai > 0")
    payload = {
        "id_global": id_global,
        # id_from van NOT NULL tren DB chua chay migration 030; ghi cung ma global.
        "id_from": id_global,
        "d_name": d_name,
        "amount": amount_vnd,
        "bank_type": bank_type,
        "stk": stk,
        "bank_name": bank_name,
        "status": WITHDRAW_STATUS_CHUA_BAO_KHACH,
        "note": "zalo-contacts payout-file apply",
    }
    res = supabase.table("withdraw_requests").insert(payload).execute()
    rows = res.data or []
    if not rows:
        raise HTTPException(status_code=500, detail="Khong tao duoc withdraw_requests")
    wid = rows[0].get("id")
    return str(wid) if wid is not None else ""


def _contact_maps_for_payout(rows: list[dict]) -> tuple[dict[str, list[dict]], dict[int, dict]]:
    """Key theo id_global (global-only)."""
    by_contact_key: dict[str, list[dict]] = {}
    by_id: dict[int, dict] = {}
    for row in rows:
        cid = row.get("id")
        if cid is None:
            continue
        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            continue
        id_g = str(row.get("id_global") or "").strip()
        if not id_g:
            continue
        by_id[cid_int] = row
        by_contact_key.setdefault(id_g, []).append(row)
    return by_contact_key, by_id


def _parse_amount_cell(v: object) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _excel_cell_to_id_global(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _resolve_payout_deduct_vnd(avail: float, amt_request: int) -> float | None:
    """
    So tien thuc te tru khoi available (float).
    None = khong du. Cho phep lech nho (lam tron / float) giua file va DB, VD 8152.8 vs 8153.
    """
    a = float(avail or 0.0)
    if amt_request <= a + 1e-6:
        return float(amt_request)
    if amt_request > a and (amt_request - a) <= _PAYOUT_FLOAT_VND_SLACK:
        return a
    return None


def _floor_vnd_to_thousand(v: float) -> int:
    """Lam trong xuong boi so 1.000 VND (don vi nghin). VD 5453 -> 5000."""
    x = max(0.0, float(v or 0.0))
    return int(math.floor(x / 1000.0 + 1e-9)) * 1000


def _payout_preview_from_dataframe(
    df: pd.DataFrame,
    *,
    by_id_global: dict[str, list[dict]],
) -> tuple[list[dict], dict[str, int]]:
    columns = [str(c).strip() for c in df.columns]
    col_id = _pick_id_global_column(columns)
    col_amt = _pick_amount_column(columns)
    col_st = _pick_status_column(columns)
    if not col_id and not col_amt:
        raise HTTPException(
            status_code=400,
            detail="File thieu cot id_global va cot So tien chuyen",
        )
    if not col_id:
        raise HTTPException(
            status_code=400,
            detail="File thieu cot id_global",
        )
    if not col_amt:
        raise HTTPException(
            status_code=400,
            detail="File thieu cot So tien chuyen",
        )
    if not col_st:
        raise HTTPException(
            status_code=400,
            detail="File thieu cot Hoan Tien chua (hoac Trang thai)",
        )

    out_rows: list[dict] = []
    summary = {
        "matched": 0,
        "skipped": 0,
        "not_found": 0,
        "ambiguous": 0,
        "insufficient": 0,
        "bad_row": 0,
        "withdraw_blocked": 0,
    }

    for i, row in df.iterrows():
        sheet_row = int(i) + 2
        id_global = _excel_cell_to_id_global(row.get(col_id))
        amt = _parse_amount_cell(row.get(col_amt))
        st_cell = row.get(col_st)

        base: dict = {
            "sheet_row": sheet_row,
            "id_global": id_global or None,
            "amount_file": None,
            "status_raw": None if st_cell is None or (isinstance(st_cell, float) and pd.isna(st_cell)) else str(st_cell),
            "result": "",
            "message": None,
            "contact_id": None,
            "d_name_db": None,
            "available_before": None,
            "available_after": None,
            "received_before": None,
            "received_after": None,
            "deduct_applied": None,
            "is_admin": False,
        }

        if not id_global:
            if amt is None and (st_cell is None or (isinstance(st_cell, float) and pd.isna(st_cell))):
                continue
            base["result"] = "bad_row"
            base["message"] = "Thieu id_global"
            summary["bad_row"] += 1
            out_rows.append(base)
            continue

        if not _is_payout_paid_cell(st_cell):
            base["result"] = "skipped"
            base["message"] = "Khong phai Da Tra Hoa Hong — bo qua"
            summary["skipped"] += 1
            out_rows.append(base)
            continue

        matches = by_id_global.get(id_global, [])
        if not matches:
            base["result"] = "not_found"
            base["message"] = "Ma khong thuoc danh sach cua ban (id_global)"
            summary["not_found"] += 1
            out_rows.append(base)
            continue
        if len(matches) > 1:
            base["result"] = "ambiguous"
            base["message"] = "Trung ma tren nhieu ban ghi — can xu ly thu cong"
            summary["ambiguous"] += 1
            out_rows.append(base)
            continue

        c = matches[0]
        cid = int(c["id"])
        avail = float(c.get("available_amount") or 0)
        recv = float(c.get("received") or 0)
        base["is_admin"] = _is_admin_zalo_contact_role(c.get("role"))

        if amt is None or amt <= 0:
            base["result"] = "bad_row"
            base["message"] = "So tien chuyen khong hop le"
            summary["bad_row"] += 1
            out_rows.append(base)
            continue
        base["amount_file"] = int(round(amt))
        amt_i = int(round(amt))
        deduct = _payout_deduct_vnd_for_contact(avail=avail, amount_vnd=amt_i)
        if deduct is None:
            base["result"] = "insufficient"
            base["message"] = f"available_amount ({avail}) < so tien chuyen ({amt_i})"
            summary["insufficient"] += 1
            out_rows.append(base)
            continue

        base["result"] = "matched"
        base["contact_id"] = cid
        base["d_name_db"] = str(c.get("d_name") or "").strip() or None
        base["deduct_applied"] = round(deduct, 4)
        base["message"] = (
            f"received += So tien chuyen {amt_i} VND "
            f"(available {round(avail, 4)} -> {round(avail - deduct, 4)})"
        )
        if abs(float(amt_i) - deduct) > 1e-6:
            base["message"] += (
                f"; tru thuc te {round(deduct, 4)} VND "
                f"(chenh lech float/lam tron <= {_PAYOUT_FLOAT_VND_SLACK} VND)"
            )
        base["available_before"] = round(avail, 4)
        base["available_after"] = round(avail - deduct, 4)
        recv_delta = int(round(float(deduct)))
        base["received_before"] = round(recv, 4)
        base["received_after"] = round(float(recv) + float(recv_delta), 4)
        amount_apply = int(round(float(deduct)))
        elig, block = _payout_withdraw_eligible(c)
        if not elig:
            base["result"] = "withdraw_blocked"
            prev = base.get("message") or "So du hop le"
            base["message"] = f"{prev}; khong chot duoc: {block}"
            summary["withdraw_blocked"] += 1
        else:
            prev = base.get("message")
            suffix = "; + withdraw_requests CHUA_BAO_KHACH khi chot"
            base["message"] = (prev + suffix) if prev else suffix.strip("; ")
            summary["matched"] += 1
        out_rows.append(base)

    return out_rows, summary


class PayoutApplyItem(BaseModel):
    contact_id: int = Field(..., gt=0)
    amount_vnd: int = Field(..., gt=0)


class PayoutApplyBody(BaseModel):
    rows: list[PayoutApplyItem] = Field(..., min_length=1, max_length=500)


@router.post("/zalo-contacts/payout-file/preview")
async def preview_zalo_payout_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Doc Excel (dong dau = header): id_global, So tien chuyen, Hoan Tien chua = Da Tra Hoa Hong."""
    try:
        raw = await file.read()
        if len(raw) > PAYOUT_IMPORT_MAX_BYTES:
            raise HTTPException(status_code=400, detail="File qua lon (toi da 5MB)")
        if not raw:
            raise HTTPException(status_code=400, detail="File rong")

        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )

        try:
            df = pd.read_excel(BytesIO(raw), header=0, engine="openpyxl")
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Khong doc duoc Excel: {exc}") from exc

        contacts = _all_zalo_contact_rows_for_owner(supabase, user_owner_global=user_owner)
        by_id_global, _by_id = _contact_maps_for_payout(contacts)
        rows, summary = _payout_preview_from_dataframe(df, by_id_global=by_id_global)

        apply_rows = [
            {
                "contact_id": r["contact_id"],
                "amount_vnd": int(round(float(r["deduct_applied"]))),
            }
            for r in rows
            if r.get("result") == "matched"
            and r.get("contact_id") is not None
            and r.get("deduct_applied") is not None
            and float(r["deduct_applied"]) > 0
        ]
        return {"rows": rows, "summary": summary, "apply_rows": apply_rows}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Payout preview loi: {exc}",
        ) from exc


@router.post("/zalo-contacts/payout-file/apply")
def apply_zalo_payout_import(
    body: PayoutApplyBody,
    current_user: dict = Depends(get_current_user),
):
    """Tru available_amount, cong received; ghi withdraw_requests status CHUA_BAO_KHACH."""
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )

        contacts = _all_zalo_contact_rows_for_owner(supabase, user_owner_global=user_owner)
        _, by_id = _contact_maps_for_payout(contacts)

        seen: set[int] = set()
        for it in body.rows:
            if it.contact_id in seen:
                raise HTTPException(status_code=400, detail="Trung contact_id trong payload")
            seen.add(it.contact_id)

        applied: list[dict] = []
        for it in body.rows:
            row = by_id.get(it.contact_id)
            if not row:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} khong ton tai hoac khong thuoc ban",
                )
            amt = int(it.amount_vnd)
            if amt <= 0:
                raise HTTPException(status_code=400, detail="amount_vnd phai > 0")

            snap = (
                supabase.table("zalo_contacts")
                .select("available_amount,received")
                .eq("id", it.contact_id)
                .limit(1)
                .execute()
            )
            snap_rows = snap.data or []
            if not snap_rows:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} khong doc duoc snapshot",
                )
            cur = snap_rows[0]
            avail = float(cur.get("available_amount") or 0)
            recv = float(cur.get("received") or 0)
            is_admin = _is_admin_zalo_contact_role(row.get("role"))
            deduct = _payout_deduct_vnd_for_contact(avail=avail, amount_vnd=amt)
            if deduct is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"contact_id={it.contact_id} id_global={row.get('id_global')} "
                        f"khong du available_amount (con {avail}, can So tien chuyen {amt})"
                    ),
                )

            amount_apply = int(round(float(deduct)))
            elig, block = _payout_withdraw_eligible(row)
            if not elig:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"contact_id={it.contact_id} id_global={row.get('id_global')} "
                        f"khong du dieu kien withdraw_requests: {block}"
                    ),
                )
            bank_fields = _withdraw_bank_fields_from_contact(row)
            assert bank_fields is not None
            bank_type, stk, bank_name = bank_fields
            id_global = str(row.get("id_global") or "").strip()
            if not id_global:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} thieu id_global",
                )

            new_avail = round(avail - float(deduct), 4)
            recv_delta = int(round(float(deduct)))
            new_recv = int(round(float(recv) + float(recv_delta)))

            res = (
                supabase.table("zalo_contacts")
                .update(
                    {
                        "available_amount": new_avail,
                        "received": new_recv,
                    }
                )
                .eq("id", it.contact_id)
                .execute()
            )
            updated = res.data or []
            if not updated:
                raise HTTPException(
                    status_code=409,
                    detail=f"Khong cap nhat duoc contact_id={it.contact_id} (so du da doi hoac bi khac xu ly)",
                )
            withdraw_id = _insert_withdraw_request_chua_bao_khach(
                supabase,
                id_global=id_global,
                d_name=str(row.get("d_name") or "").strip() or None,
                amount_vnd=amount_apply,
                bank_type=bank_type,
                stk=stk,
                bank_name=bank_name,
            )
            applied.append(
                {
                    "contact_id": it.contact_id,
                    "id_global": str(row.get("id_global") or "").strip() or None,
                    "amount_vnd": amt,
                    "is_admin": is_admin,
                    "deduct_applied": round(float(deduct), 4),
                    "available_after": new_avail,
                    "received_after": new_recv,
                    "withdraw_request_id": withdraw_id,
                    "withdraw_status": WITHDRAW_STATUS_CHUA_BAO_KHACH,
                }
            )
            row["available_amount"] = new_avail
            row["received"] = new_recv

        return {"applied_count": len(applied), "applied": applied}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Payout apply loi: {exc}",
        ) from exc


@router.get("/orders")
def list_commission_orders(
    limit: int = Query(default=200, ge=1, le=1000),
    placed_within_days: int | None = Query(
        default=None,
        description=(
            "Loc order_placed_at: tu 00:00 VN cua (hom_nay_lich_VN - N) den hien tai (1, 3, 7, 14). Bo qua = tat ca."
        ),
    ),
):
    """Doc ban ghi da import trong affiliate_commission_orders (moi nhat truoc)."""
    if placed_within_days is not None and placed_within_days not in PLACED_WITHIN_DAYS_ALLOWED:
        raise HTTPException(
            status_code=400,
            detail="placed_within_days chi chap nhan 1, 3, 7 hoac 14",
        )
    try:
        supabase = get_supabase_client()
        query = supabase.table("affiliate_commission_orders").select("*")
        if placed_within_days is not None:
            start_utc, end_utc = _placed_within_vn_calendar_to_now(placed_within_days)
            query = query.gte("order_placed_at", start_utc.isoformat()).lte(
                "order_placed_at", end_utc.isoformat()
            )
        result = query.order("order_placed_at", desc=True).limit(limit).execute()
        items = result.data or []
        return {"count": len(items), "items": items}
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query Supabase loi: {exc}",
        ) from exc


@router.get("/payout-sync-logs")
def list_payout_sync_logs(
    limit: int = Query(default=500, ge=1, le=2000),
    placed_within_days: int | None = Query(
        default=None,
        description=(
            "Loc created_at: tu 00:00 VN cua (hom_nay_lich_VN - N) den hien tai (1, 3, 7, 14). Bo qua = tat ca."
        ),
    ),
):
    """Doc lich su dong bo tien (commission_payout_sync_log), moi nhat truoc."""
    if placed_within_days is not None and placed_within_days not in PLACED_WITHIN_DAYS_ALLOWED:
        raise HTTPException(
            status_code=400,
            detail="placed_within_days chi chap nhan 1, 3, 7 hoac 14",
        )
    try:
        supabase = get_supabase_client()
        query = supabase.table("commission_payout_sync_log").select("*")
        if placed_within_days is not None:
            start_utc, end_utc = _placed_within_vn_calendar_to_now(placed_within_days)
            query = query.gte("created_at", start_utc.isoformat()).lte(
                "created_at", end_utc.isoformat()
            )
        result = query.order("created_at", desc=True).limit(limit).execute()
        items = result.data or []
        return {"count": len(items), "items": items}
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query payout sync log loi: {exc}",
        ) from exc


@router.get("/order-splits")
def list_commission_order_splits(
    limit: int = Query(default=500, ge=1, le=2000),
    order_id: str | None = Query(
        default=None,
        description="Loc theo order_id (exact). Bo qua = tat ca.",
    ),
    placed_within_days: int | None = Query(
        default=None,
        description=(
            "Loc created_at: tu 00:00 VN cua (hom_nay_lich_VN - N) den hien tai (1, 3, 7, 14). Bo qua = tat ca."
        ),
    ),
):
    """Doc affiliate_commission_order_splits (phan tang), moi nhat truoc."""
    if placed_within_days is not None and placed_within_days not in PLACED_WITHIN_DAYS_ALLOWED:
        raise HTTPException(
            status_code=400,
            detail="placed_within_days chi chap nhan 1, 3, 7 hoac 14",
        )
    try:
        supabase = get_supabase_client()
        query = supabase.table(SPLIT_TABLE).select("*")
        if order_id is not None and str(order_id).strip():
            query = query.eq("order_id", str(order_id).strip())
        if placed_within_days is not None:
            start_utc, end_utc = _placed_within_vn_calendar_to_now(placed_within_days)
            query = query.gte("created_at", start_utc.isoformat()).lte(
                "created_at", end_utc.isoformat()
            )
        result = query.order("created_at", desc=True).limit(limit).execute()
        items = result.data or []
        id_g_values = sorted(
            {
                str(row.get("id_global") or "").strip()
                for row in items
                if str(row.get("id_global") or "").strip()
            }
        )
        contact_map: dict[str, dict[str, str | None]] = {}
        if id_g_values:
            contact_map.update(_build_contact_map_by_global(supabase, id_g_values))
        if contact_map:
            for row in items:
                k = str(row.get("id_global") or "").strip()
                info = contact_map.get(k) if k else None
                row["d_name"] = (info or {}).get("name")
        else:
            for row in items:
                row["d_name"] = None
        return {"count": len(items), "items": items}
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query order splits loi: {exc}",
        ) from exc


def _unwrap_rpc_jsonb(result) -> dict:
    """Chuan hoa response RPC tra ve jsonb (dict / list mot phan tu / key ten ham)."""
    data = result.data
    if data is None:
        return {}
    if isinstance(data, dict):
        if "orders_updated" in data:
            return data
        if len(data) == 1:
            inner = next(iter(data.values()))
            if isinstance(inner, dict):
                return inner
        return data
    if isinstance(data, list) and len(data) == 1 and isinstance(data[0], dict):
        row = data[0]
        if "orders_updated" in row:
            return row
        if len(row) == 1:
            inner = next(iter(row.values()))
            if isinstance(inner, dict):
                return inner
    return {}


def _analyze_import_rows(content: bytes, filename: str) -> dict:
    try:
        rows = parse_and_aggregate_report(content, filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Doc file loi: {exc}") from exc

    if not rows:
        return {
            "rows_to_upsert": [],
            "skipped_already_paid": 0,
            "unique_orders_from_file": 0,
            "sub_id_values": [],
            "convert_info_map": {},
            "matched_convert": 0,
            "matched_contact": 0,
            "matched_config": 0,
            "missing_config": 0,
            "hoa_hong": {
                "ok": False,
                "value_1": 0.0,
                "value_2": 0.0,
                "value_3": 0.0,
                "value_4": 0.0,
            },
        }

    supabase = get_supabase_client()
    for row in rows:
        row["source_filename"] = filename
        row["id_global"] = None
        row["name"] = None
        row["hh_user"] = 0

    sub_id_values = sorted(
        {
            str(row.get("sub_id1") or "").strip()
            for row in rows
            if str(row.get("sub_id1") or "").strip()
        }
    )

    try:
        convert_info_map = _build_convert_info_map(supabase, sub_id_values)
        id_gz_values = sorted(
            {
                str(info.get("id_globalzalo") or "").strip()
                for info in convert_info_map.values()
                if str(info.get("id_globalzalo") or "").strip()
            }
        )
        contact_map = _build_contact_map_by_global(supabase, id_gz_values)
        h1, h2, h3, h4, hoa_hong_ok = _fetch_hoa_hong_pct_quads(supabase)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Tra cuu convert_results/zalo_contacts loi: {exc}",
        ) from exc

    matched_convert = 0
    matched_contact = 0
    matched_config = 0
    missing_config = 0
    for row in rows:
        sub_id1 = str(row.get("sub_id1") or "").strip()
        if not sub_id1:
            continue
        convert_info = convert_info_map.get(sub_id1)
        if not convert_info:
            continue
        matched_convert += 1
        id_gz = str(convert_info.get("id_globalzalo") or "").strip()
        if not id_gz:
            continue
        contact = contact_map.get(id_gz)
        if not contact:
            continue

        if hoa_hong_ok:
            matched_config += 1
        else:
            missing_config += 1

        net_commission = float(row.get("net_affiliate_commission") or 0)
        id_global_c = str((contact or {}).get("id_global") or "").strip() or id_gz or None
        row["id_global"] = id_global_c
        row["name"] = (contact or {}).get("name")
        _, _, hh_u = _amounts_from_hoa_hong_net(net_commission, h1, h2, h3, h4)
        row["hh_user"] = hh_u if hoa_hong_ok else 0.0
        matched_contact += 1

    order_id_list = [str(r.get("order_id") or "") for r in rows]
    try:
        existing_order_status = _fetch_existing_order_status_by_order_id(supabase, order_id_list)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Doc order_status hien co loi: {exc}",
        ) from exc

    skipped_already_paid = 0
    rows_to_upsert: list[dict] = []
    for r in rows:
        oid = str(r.get("order_id") or "").strip()
        prev = existing_order_status.get(oid)
        if str(prev or "").strip() == ORDER_STATUS_PAID_SYNCED:
            skipped_already_paid += 1
        else:
            rows_to_upsert.append(r)
    _apply_order_status_transition(rows_to_upsert, existing_order_status)

    return {
        "rows_to_upsert": rows_to_upsert,
        "skipped_already_paid": skipped_already_paid,
        "unique_orders_from_file": len(rows),
        "sub_id_values": sub_id_values,
        "convert_info_map": convert_info_map,
        "matched_convert": matched_convert,
        "matched_contact": matched_contact,
        "matched_config": matched_config,
        "missing_config": missing_config,
        "hoa_hong": {
            "ok": hoa_hong_ok,
            "value_1": h1,
            "value_2": h2,
            "value_3": h3,
            "value_4": h4,
        },
    }


def _preview_item_sort_key(item: dict) -> tuple[int, str]:
    """Uu tien hien thi: loi split -> Da huy -> doi trang thai -> se cap nhat -> khong doi."""
    if item.get("split_issue"):
        rank = 0
    elif _is_order_status_cancelled(item.get("order_status")):
        rank = 1
    elif item.get("order_status_transition"):
        rank = 2
    elif item.get("is_unchanged"):
        rank = 4
    else:
        rank = 3
    return (rank, str(item.get("order_id") or ""))


def _build_import_preview_items(
    supabase,
    rows_to_upsert: list[dict],
    convert_info_map: dict[str, dict[str, str | None]],
    *,
    existing_by_order_id: dict[str, dict] | None = None,
    hoa_hong: dict | None = None,
) -> list[dict]:
    out: list[dict] = []
    group_cache: dict[str, tuple[str | None, str | None, str | None]] = {}
    agency_globals_for_lookup: set[str] = set()
    hh = hoa_hong or {}
    v1 = float(hh.get("value_1") or 0)
    v2 = float(hh.get("value_2") or 0)
    v3 = float(hh.get("value_3") or 0)
    v4 = float(hh.get("value_4") or 0)
    hoa_ok = bool(hh.get("ok"))
    for row in rows_to_upsert:
        sub_id1 = str(row.get("sub_id1") or "").strip()
        convert_info = convert_info_map.get(sub_id1) if sub_id1 else None
        id_globalgroup = str((convert_info or {}).get("id_globalgroup") or "").strip() or None

        agency_id_global = None
        owner_id_global_main = None
        group_id_global = None
        split_issue = None
        if not sub_id1:
            split_issue = "thieu_sub_id1"
        elif not convert_info:
            split_issue = "thieu_convert_results"
        elif not str((convert_info or {}).get("id_globalzalo") or "").strip():
            split_issue = "thieu_id_globalzalo"
        elif not id_globalgroup:
            split_issue = "thieu_id_globalgroup"
        else:
            if id_globalgroup not in group_cache:
                group_cache[id_globalgroup] = _fetch_group_globals_for_import(
                    supabase, id_globalgroup
                )
            group_id_global, agency_id_global, owner_id_global_main = group_cache[id_globalgroup]
            if not owner_id_global_main:
                split_issue = "thieu_id_global_main"
            elif not agency_id_global:
                split_issue = "thieu_id_globalzalo_nhom"
        if agency_id_global:
            agency_globals_for_lookup.add(agency_id_global)

        net = float(row.get("net_affiliate_commission") or 0)
        oid = str(row.get("order_id") or "").strip()
        db_row = (existing_by_order_id or {}).get(oid) if oid else None
        is_unchanged = (
            db_row is not None and _incoming_commission_row_equals_db(row, db_row)
        )
        agency_amount, owner_amount, _hh_calc = _amounts_from_hoa_hong_net(net, v1, v2, v3, v4)
        if not hoa_ok:
            agency_amount = 0.0
            owner_amount = 0.0
        out.append(
            {
                "order_id": oid,
                "order_status": str(row.get("order_status") or ""),
                "order_status_transition": row.get("order_status_transition"),
                "is_unchanged": is_unchanged,
                "net_affiliate_commission": net,
                "id_global": row.get("id_global"),
                "name": row.get("name"),
                "hh_user": row.get("hh_user"),
                "id_globalgroup": id_globalgroup,
                "group_id_global": group_id_global or id_globalgroup,
                "agency_id_global": agency_id_global,
                "agency_name": None,
                "owner_id_global_main": owner_id_global_main,
                "agency_amount": agency_amount,
                "owner_amount": owner_amount,
                "split_issue": split_issue,
            }
        )
    if agency_globals_for_lookup:
        agency_contact_map = _build_contact_map_by_global(
            supabase, sorted(agency_globals_for_lookup)
        )
        for item in out:
            ag = str(item.get("agency_id_global") or "").strip()
            item["agency_name"] = (agency_contact_map.get(ag) or {}).get("name") if ag else None
    out.sort(key=_preview_item_sort_key)
    return out


@router.post("/sync-hh-to-zalo-bill-preview")
async def sync_hh_to_zalo_bill_preview(file: UploadFile = File(...)):
    """
    Doc file Bill Conversion (Shopee): don co Trang thai dat hang = Hoan thanh trong file.
    Tra ve preview doi chieu DB (khong ghi DB).
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Thieu ten file")
    suffix = file.filename.lower()
    if not (suffix.endswith(".csv") or suffix.endswith(".xlsx") or suffix.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Chi ho tro file .csv, .xlsx hoac .xls")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File rong")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File vuot qua 15MB")

    try:
        parsed = parse_bill_conversion_completed_order_ids(content, file.filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Doc file loi: {exc}") from exc

    order_ids: list[str] = parsed["order_ids"]
    supabase = get_supabase_client()
    preview_items = _build_bill_sync_hh_preview(supabase, order_ids)
    eligible_n = sum(1 for it in preview_items if it.get("eligible"))
    return {
        "ok": True,
        "source_filename": file.filename,
        "rows_total_in_file": parsed["rows_total_in_file"],
        "rows_completed_in_file": parsed["rows_completed_in_file"],
        "unique_orders_completed": parsed["unique_orders_completed"],
        "preview_eligible_count": eligible_n,
        "preview_skip_count": len(preview_items) - eligible_n,
        "preview_items": preview_items,
    }


@router.post("/sync-hh-to-zalo-bill-apply")
async def sync_hh_to_zalo_bill_apply(file: UploadFile = File(...)):
    """
    Cung dinh dang file nhu preview: parse lai order_id Hoan thanh, goi RPC chi dong bo cac don do.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Thieu ten file")
    suffix = file.filename.lower()
    if not (suffix.endswith(".csv") or suffix.endswith(".xlsx") or suffix.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Chi ho tro file .csv, .xlsx hoac .xls")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File rong")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File vuot qua 15MB")

    try:
        parsed = parse_bill_conversion_completed_order_ids(content, file.filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Doc file loi: {exc}") from exc

    order_ids: list[str] = parsed["order_ids"]
    if not order_ids:
        raise HTTPException(
            status_code=400,
            detail="File khong co dong Trang thai dat hang = Hoan thanh hop le",
        )

    supabase = get_supabase_client()
    preview_items = _build_bill_sync_hh_preview(supabase, order_ids)
    eligible_n = sum(1 for it in preview_items if it.get("eligible"))
    if eligible_n == 0:
        raise HTTPException(
            status_code=400,
            detail="Khong co don nao du dieu kien dong bo (DB Hoan thanh + id_global + zalo_contacts)",
        )

    try:
        result = supabase.rpc(
            "sync_commission_hh_to_zalo",
            {"p_restrict_order_ids": order_ids},
        ).execute()
    except Exception as exc:
        logger.exception("[commission-sync] RPC bill apply failed")
        raise HTTPException(status_code=500, detail=f"RPC loi: {exc}") from exc

    payload = _unwrap_rpc_jsonb(result)
    if not payload:
        raise HTTPException(status_code=500, detail="RPC tra ve rong")
    logger.info(
        "[commission-sync-bill] orders_updated=%s skipped=%s total=%s file=%s",
        payload.get("orders_updated"),
        payload.get("orders_skipped_no_contact"),
        payload.get("total_amount_added"),
        file.filename,
    )
    return {
        **payload,
        "source_filename": file.filename,
        "restrict_order_count": len(order_ids),
        "preview_eligible_count": eligible_n,
    }


@router.post("/sync-hh-to-zalo")
def sync_hh_to_zalo():
    """
    Goi RPC sync_commission_hh_to_zalo (khong loc order_id): tat ca don Hoan thanh du dieu kien.
    """
    try:
        supabase = get_supabase_client()
        result = supabase.rpc("sync_commission_hh_to_zalo", {}).execute()
    except Exception as exc:
        logger.exception("[commission-sync] RPC failed")
        raise HTTPException(status_code=500, detail=f"RPC loi: {exc}") from exc

    payload = _unwrap_rpc_jsonb(result)
    if not payload:
        raise HTTPException(status_code=500, detail="RPC tra ve rong")
    logger.info(
        "[commission-sync] orders_updated=%s skipped=%s contacts=%s total=%s",
        payload.get("orders_updated"),
        payload.get("orders_skipped_no_contact"),
        len(payload.get("contacts") or []) if isinstance(payload.get("contacts"), list) else 0,
        payload.get("total_amount_added"),
    )
    return payload


@router.post("/import-preview")
async def import_commission_report_preview(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Thieu ten file")
    suffix = file.filename.lower()
    if not (suffix.endswith(".csv") or suffix.endswith(".xlsx") or suffix.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Chi ho tro file .csv, .xlsx hoac .xls")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File rong")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File vuot qua 15MB")

    analyzed = _analyze_import_rows(content, file.filename)
    supabase = get_supabase_client()
    order_ids_preview = [
        str(r.get("order_id") or "").strip()
        for r in analyzed["rows_to_upsert"]
        if str(r.get("order_id") or "").strip()
    ]
    existing_map = _fetch_orders_map_by_order_id(supabase, order_ids_preview)
    rows_for_upsert, would_skip_unchanged = _partition_rows_skip_unchanged(
        supabase, analyzed["rows_to_upsert"]
    )
    preview_items = _build_import_preview_items(
        supabase,
        analyzed["rows_to_upsert"],
        analyzed["convert_info_map"],
        existing_by_order_id=existing_map,
        hoa_hong=analyzed.get("hoa_hong"),
    )
    preview_with_issue = sum(1 for item in preview_items if item.get("split_issue"))
    preview_counts = {
        "total": len(preview_items),
        "issue": preview_with_issue,
        "cancelled": sum(
            1 for item in preview_items if _is_order_status_cancelled(item.get("order_status"))
        ),
        "transition": sum(1 for item in preview_items if item.get("order_status_transition")),
        "unchanged": sum(1 for item in preview_items if item.get("is_unchanged")),
    }
    return {
        "ok": True,
        "source_filename": file.filename,
        "unique_orders": analyzed["unique_orders_from_file"],
        "would_upsert": len(rows_for_upsert),
        "would_skip_unchanged": would_skip_unchanged,
        "skipped_already_paid": analyzed["skipped_already_paid"],
        "preview_items": preview_items,
        "preview_with_issue": preview_with_issue,
        "preview_counts": preview_counts,
        "lookup": {
            "sub_id1_count": len(analyzed["sub_id_values"]),
            "matched_convert_results": analyzed["matched_convert"],
            "missing_convert_results": sum(
                1 for s in analyzed["sub_id_values"] if s not in analyzed["convert_info_map"]
            ),
            "matched_zalo_contacts": analyzed["matched_contact"],
            "missing_zalo_contacts": max(
                analyzed["matched_convert"] - analyzed["matched_contact"], 0
            ),
            "matched_commission_config": analyzed["matched_config"],
            "missing_commission_config": analyzed["missing_config"],
            "hoa_hong": analyzed.get("hoa_hong"),
        },
    }


@router.post("/import")
async def import_commission_report(file: UploadFile = File(...)):
    """
    Nhan file CSV/XLSX bao cao hoa hong Shopee, gop theo ID don hang, upsert vao Supabase.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Thieu ten file")
    logger.info("[commission-import] start filename=%s", file.filename)

    suffix = file.filename.lower()
    if not (suffix.endswith(".csv") or suffix.endswith(".xlsx") or suffix.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Chi ho tro file .csv, .xlsx hoac .xls")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File rong")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File vuot qua 15MB")

    analyzed = _analyze_import_rows(content, file.filename)
    rows = analyzed["rows_to_upsert"]
    skipped_already_paid_count = analyzed["skipped_already_paid"]
    unique_orders_from_file = analyzed["unique_orders_from_file"]
    sub_id_values = analyzed["sub_id_values"]
    convert_info_map = analyzed["convert_info_map"]
    matched_convert = analyzed["matched_convert"]
    matched_contact = analyzed["matched_contact"]
    matched_config = analyzed["matched_config"]
    missing_config = analyzed["missing_config"]

    if unique_orders_from_file == 0:
        return {
            "upserted": 0,
            "unique_orders": 0,
            "skipped_already_paid": 0,
            "skipped_unchanged": 0,
            "import_batch_id": None,
            "split_sync": {
                "splits_rows_upserted": 0,
                "orders_splits_status_only": 0,
                "splits_rows_deleted": 0,
            },
            "message": "Khong co dong hop le sau khi doc file",
        }

    supabase = get_supabase_client()
    logger.info(
        "[commission-import] skipped_already_paid=%s before_filter_upsert=%s",
        skipped_already_paid_count,
        len(rows),
    )

    rows, skipped_unchanged = _partition_rows_skip_unchanged(supabase, rows)
    logger.info(
        "[commission-import] after_skip_unchanged remaining_for_upsert=%s skipped_unchanged=%s",
        len(rows),
        skipped_unchanged,
    )

    if not rows:
        return {
            "upserted": 0,
            "unique_orders": unique_orders_from_file,
            "skipped_already_paid": skipped_already_paid_count,
            "skipped_unchanged": skipped_unchanged,
            "source_filename": file.filename,
            "import_batch_id": None,
            "split_sync": {
                "splits_rows_upserted": 0,
                "orders_splits_status_only": 0,
                "splits_rows_deleted": 0,
            },
            "lookup": {
                "sub_id1_count": len(sub_id_values),
                "matched_convert_results": matched_convert,
                "missing_convert_results": sum(
                    1 for s in sub_id_values if s not in convert_info_map
                ),
                "matched_zalo_contacts": matched_contact,
                "missing_zalo_contacts": max(matched_convert - matched_contact, 0),
                "matched_commission_config": matched_config,
                "missing_commission_config": missing_config,
                "hoa_hong": analyzed.get("hoa_hong"),
            },
        }

    import_batch_id = str(uuid.uuid4())
    upserted = 0
    for start in range(0, len(rows), UPSERT_CHUNK):
        chunk = [_strip_legacy_import_order_fields(r) for r in rows[start : start + UPSERT_CHUNK]]
        try:
            supabase.table("affiliate_commission_orders").upsert(
                chunk,
                on_conflict="order_id",
            ).execute()
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=f"Luu Supabase loi: {exc}",
            ) from exc
        upserted += len(chunk)

    split_stats = {
        "splits_rows_upserted": 0,
        "orders_splits_status_only": 0,
        "splits_rows_deleted": 0,
    }
    order_ids_for_splits = sorted(
        {str(r.get("order_id") or "").strip() for r in rows if str(r.get("order_id") or "").strip()}
    )
    if order_ids_for_splits:
        try:
            orders_from_db = _fetch_commission_orders_rows_by_order_ids(
                supabase, order_ids_for_splits
            )
            split_stats = _sync_commission_order_splits_for_import(
                supabase,
                orders_from_db,
                import_batch_id,
                convert_info_map,
                hoa_v1=float((analyzed.get("hoa_hong") or {}).get("value_1") or 0),
                hoa_v2=float((analyzed.get("hoa_hong") or {}).get("value_2") or 0),
                hoa_v4=float((analyzed.get("hoa_hong") or {}).get("value_4") or 0),
            )
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("[commission-import] splits sync failed")
            raise HTTPException(
                status_code=500,
                detail=f"Dong bo splits loi: {exc}",
            ) from exc

    return {
        "upserted": upserted,
        "unique_orders": unique_orders_from_file,
        "skipped_already_paid": skipped_already_paid_count,
        "skipped_unchanged": skipped_unchanged,
        "source_filename": file.filename,
        "import_batch_id": import_batch_id,
        "split_sync": split_stats,
        "lookup": {
            "sub_id1_count": len(sub_id_values),
            "matched_convert_results": matched_convert,
            "missing_convert_results": sum(
                1 for s in sub_id_values if s not in convert_info_map
            ),
            "matched_zalo_contacts": matched_contact,
            "missing_zalo_contacts": max(matched_convert - matched_contact, 0),
            "matched_commission_config": matched_config,
            "missing_commission_config": missing_config,
            "hoa_hong": analyzed.get("hoa_hong"),
        },
    }
