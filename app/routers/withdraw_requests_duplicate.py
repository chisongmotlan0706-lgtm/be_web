import logging
from io import BytesIO

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.auth import get_current_user, user_owner_global_zalo
from app.db import get_supabase_client
from app.routers.commission_import import (
    LOOKUP_CHUNK,
    PAYOUT_IMPORT_MAX_BYTES,
    PAYOUT_STATUS_PAID,
    STATUS_BANK_LOI,
    WITHDRAW_STATUS_CHUA_BAO_KHACH,
    _all_zalo_contact_rows_for_owner,
    _chunked,
    _contact_maps_for_payout,
    _excel_cell_to_id_global,
    _fetch_chuyen_tien_settings,
    _fetch_contact_globals_for_group_ids,
    _fetch_group_ids_by_owner,
    _fetch_order_balance_maps_by_id_zl,
    _fetch_split_balance_maps_by_id_zl,
    _floor_vnd_to_thousand,
    _fold_col_name,
    _insert_withdraw_request_chua_bao_khach,
    _is_admin_zalo_contact_role,
    _is_payout_paid_cell,
    _parse_amount_cell,
    _payout_deduct_vnd_for_contact,
    _payout_withdraw_eligible,
    _pick_amount_column,
    _pick_id_global_column,
    _pick_status_column,
    _withdraw_bank_fields_from_contact,
)

router = APIRouter(prefix="/commission-report", tags=["commission-report"])
logger = logging.getLogger(__name__)

DUPLICATE_TABLE = "withdraw_requests_duplicate"
DUPLICATE_STATUS_PENDING = "PENDING"
DUPLICATE_STATUS_COMPLETED = "COMPLETED"
PAYOUT_STATUS_LOI = "Lỗi"


def _is_payout_loi_cell(v: object) -> bool:
    return _fold_col_name(str(v if v is not None else "")) == "loi"


def _allowed_contact_globals_for_owner(supabase, user_owner_global: str) -> set[str]:
    owner = str(user_owner_global or "").strip()
    if not owner:
        return set()
    all_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=False
    )
    if not all_group_globals:
        return set()
    return _fetch_contact_globals_for_group_ids(supabase, all_group_globals)


def _fetch_pending_duplicates_for_owner(
    supabase,
    *,
    user_owner_global: str,
    limit: int,
) -> list[dict]:
    allowed = _allowed_contact_globals_for_owner(supabase, user_owner_global)
    if not allowed:
        return []

    out: list[dict] = []
    offset = 0
    page_size = 1000
    while len(out) < limit:
        result = (
            supabase.table(DUPLICATE_TABLE)
            .select("*")
            .eq("status", DUPLICATE_STATUS_PENDING)
            .order("created_at", desc=True)
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = result.data or []
        if not rows:
            break
        for row in rows:
            id_g = str(row.get("id_global") or "").strip()
            if id_g and id_g in allowed:
                out.append(row)
                if len(out) >= limit:
                    break
        if len(rows) < page_size:
            break
        offset += page_size
    return out[:limit]


def _contact_by_global_map(rows: list[dict]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in rows:
        id_g = str(row.get("id_global") or "").strip()
        if id_g and id_g not in out:
            out[id_g] = row
    return out


def _enrich_duplicate_items(
    supabase,
    *,
    duplicates: list[dict],
    user_owner_global: str,
) -> list[dict]:
    owner = str(user_owner_global or "").strip()
    if not duplicates or not owner:
        return []

    all_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=False
    )
    allowed_contact_global = _fetch_contact_globals_for_group_ids(supabase, all_group_globals)
    order_pending_map, order_completed_map = _fetch_order_balance_maps_by_id_zl(
        supabase, allowed_contact_global_ids=allowed_contact_global
    )
    active_group_globals = _fetch_group_ids_by_owner(
        supabase, owner_id_global_main=owner, active_only=True
    )
    active_member_global = (
        _fetch_contact_globals_for_group_ids(supabase, active_group_globals)
        if active_group_globals
        else set()
    )
    split_pending_map, split_completed_map = _fetch_split_balance_maps_by_id_zl(
        supabase, active_member_global
    )

    id_globals = sorted(
        {str(d.get("id_global") or "").strip() for d in duplicates if str(d.get("id_global") or "").strip()}
    )
    contact_rows: list[dict] = []
    for chunk in _chunked(id_globals, LOOKUP_CHUNK):
        res = (
            supabase.table("zalo_contacts")
            .select(
                "id,id_global,d_name,id_global_gr,available_amount,received,role,bank_name,bank_type,stk,status_bank"
            )
            .in_("id_global", chunk)
            .execute()
        )
        contact_rows.extend(res.data or [])
    contact_map = _contact_by_global_map(contact_rows)

    items: list[dict] = []
    for dup in duplicates:
        dup_id = str(dup.get("id") or "").strip()
        id_g = str(dup.get("id_global") or "").strip()
        contact = contact_map.get(id_g) or {}
        id_c = str(contact.get("id_global") or "").strip()
        order_pending = order_pending_map.get(id_c, 0.0)
        order_completed = order_completed_map.get(id_c, 0.0)
        split_pending = split_pending_map.get(id_c, 0.0)
        split_completed = split_completed_map.get(id_c, 0.0)
        avail = float(contact.get("available_amount") or dup.get("amount") or 0)
        items.append(
            {
                "duplicate_id": dup_id,
                "contact_id": contact.get("id"),
                "id_global": id_g or None,
                "id_global_gr": contact.get("id_global_gr"),
                "d_name": dup.get("d_name") or contact.get("d_name"),
                "bank_type": dup.get("bank_type") or contact.get("bank_type"),
                "bank_name": dup.get("bank_name") or contact.get("bank_name"),
                "stk": dup.get("stk") or contact.get("stk"),
                "amount": float(dup.get("amount") or 0),
                "status_bank": contact.get("status_bank"),
                "role": contact.get("role"),
                "dang_giao_hang": round(order_pending + split_pending, 4),
                "cho_duyet": round(order_completed + split_completed, 4),
                "tien_co_the_rut": float(_floor_vnd_to_thousand(avail)),
                "da_rut_ve_bank": round(float(contact.get("received") or 0), 4),
            }
        )
    return items


def _duplicate_workbook(items: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chuyen tien duplicate"

    headers = [
        "STT",
        "id_global",
        "id_global_gr",
        "Số tiền chuyển",
        "TK hưởng",
        "Tên người hưởng",
        "bank_type",
        "Nội dung",
        "Hoàn Tiền chưa",
    ]
    ws.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(items, start=1):
        id_g = str(item.get("id_global") or "").strip()
        id_gr = str(item.get("id_global_gr") or "").strip()
        bank_name = str(item.get("bank_name") or "").strip()
        bank = str(item.get("bank_type") or "").strip()
        stk_raw = str(item.get("stk") or "").strip()
        amt = _floor_vnd_to_thousand(float(item.get("amount") or item.get("tien_co_the_rut") or 0))
        noi = f"Hoan tien hoa hong {amt}"
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=idx)
        c_id = ws.cell(row=row_num, column=2, value=id_g if id_g else "")
        c_id.number_format = "@"
        c_gid = ws.cell(row=row_num, column=3, value=id_gr if id_gr else "")
        c_gid.number_format = "@"
        ws.cell(row=row_num, column=4, value=amt)
        c_stk = ws.cell(row=row_num, column=5, value=stk_raw if stk_raw else "")
        c_stk.number_format = "@"
        ws.cell(row=row_num, column=6, value=bank_name if bank_name else id_g)
        ws.cell(row=row_num, column=7, value=bank)
        ws.cell(row=row_num, column=8, value=noi)
        ws.cell(row=row_num, column=9, value="")

    widths = (6, 18, 18, 16, 18, 28, 22, 40, 22)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _pending_duplicate_map_by_global(
    supabase, *, user_owner_global: str, max_rows: int = 20000
) -> dict[str, dict]:
    dups = _fetch_pending_duplicates_for_owner(
        supabase, user_owner_global=user_owner_global, limit=max_rows
    )
    out: dict[str, dict] = {}
    for d in dups:
        id_g = str(d.get("id_global") or "").strip()
        if not id_g:
            continue
        if id_g in out:
            continue
        out[id_g] = d
    return out


def _duplicate_payout_preview_from_dataframe(
    df: pd.DataFrame,
    *,
    by_id_global: dict[str, list[dict]],
    pending_by_global: dict[str, dict],
) -> tuple[list[dict], dict[str, int]]:
    columns = [str(c).strip() for c in df.columns]
    col_id = _pick_id_global_column(columns)
    col_amt = _pick_amount_column(columns)
    col_st = _pick_status_column(columns)
    if not col_id:
        raise HTTPException(status_code=400, detail="File thieu cot id_global")
    if not col_amt:
        raise HTTPException(status_code=400, detail="File thieu cot So tien chuyen")
    if not col_st:
        raise HTTPException(
            status_code=400,
            detail="File thieu cot Hoan Tien chua (hoac Trang thai)",
        )

    out_rows: list[dict] = []
    summary = {
        "matched": 0,
        "loi_matched": 0,
        "skipped": 0,
        "not_found": 0,
        "ambiguous": 0,
        "insufficient": 0,
        "bad_row": 0,
        "withdraw_blocked": 0,
        "no_pending_duplicate": 0,
    }

    for i, row in df.iterrows():
        sheet_row = int(i) + 2
        id_global = _excel_cell_to_id_global(row.get(col_id))
        amt = _parse_amount_cell(row.get(col_amt))
        st_cell = row.get(col_st)
        is_paid = _is_payout_paid_cell(st_cell)
        is_loi = _is_payout_loi_cell(st_cell)

        base: dict = {
            "sheet_row": sheet_row,
            "id_global": id_global or None,
            "amount_file": None,
            "status_raw": None
            if st_cell is None or (isinstance(st_cell, float) and pd.isna(st_cell))
            else str(st_cell),
            "result": "",
            "message": None,
            "contact_id": None,
            "duplicate_id": None,
            "d_name_db": None,
            "available_before": None,
            "available_after": None,
            "received_before": None,
            "received_after": None,
            "deduct_applied": None,
            "is_admin": False,
        }

        if not id_global:
            if amt is None and (st_cell is None or (isinstance(st_cell, float) and pd.isna(st_cell))):
                continue
            base["result"] = "bad_row"
            base["message"] = "Thieu id_global"
            summary["bad_row"] += 1
            out_rows.append(base)
            continue

        if not is_paid and not is_loi:
            base["result"] = "skipped"
            base["message"] = f"Khong phai {PAYOUT_STATUS_PAID} hoac {PAYOUT_STATUS_LOI} — bo qua"
            summary["skipped"] += 1
            out_rows.append(base)
            continue

        dup = pending_by_global.get(id_global)
        if not dup:
            base["result"] = "no_pending_duplicate"
            base["message"] = "Khong co withdraw_requests_duplicate PENDING cho id_global nay"
            summary["no_pending_duplicate"] += 1
            out_rows.append(base)
            continue

        base["duplicate_id"] = str(dup.get("id") or "").strip() or None

        matches = by_id_global.get(id_global, [])
        if not matches:
            base["result"] = "not_found"
            base["message"] = "Ma khong thuoc danh sach cua ban (id_global)"
            summary["not_found"] += 1
            out_rows.append(base)
            continue
        if len(matches) > 1:
            base["result"] = "ambiguous"
            base["message"] = "Trung ma tren nhieu ban ghi — can xu ly thu cong"
            summary["ambiguous"] += 1
            out_rows.append(base)
            continue

        c = matches[0]
        cid = int(c["id"])
        base["contact_id"] = cid
        base["d_name_db"] = str(c.get("d_name") or "").strip() or None
        base["is_admin"] = _is_admin_zalo_contact_role(c.get("role"))

        if is_loi:
            base["result"] = "loi_matched"
            base["message"] = f"Danh dau status_bank = {STATUS_BANK_LOI}; duplicate -> COMPLETED khi chot"
            summary["loi_matched"] += 1
            out_rows.append(base)
            continue

        avail = float(c.get("available_amount") or 0)
        recv = float(c.get("received") or 0)
        if amt is None or amt <= 0:
            base["result"] = "bad_row"
            base["message"] = "So tien chuyen khong hop le"
            summary["bad_row"] += 1
            out_rows.append(base)
            continue
        base["amount_file"] = int(round(amt))
        amt_i = int(round(amt))
        deduct = _payout_deduct_vnd_for_contact(avail=avail, amount_vnd=amt_i)
        if deduct is None:
            base["result"] = "insufficient"
            base["message"] = f"available_amount ({avail}) < so tien chuyen ({amt_i})"
            summary["insufficient"] += 1
            out_rows.append(base)
            continue

        base["result"] = "matched"
        base["deduct_applied"] = round(deduct, 4)
        base["message"] = (
            f"received += So tien chuyen {amt_i} VND "
            f"(available {round(avail, 4)} -> {round(avail - deduct, 4)})"
        )
        base["available_before"] = round(avail, 4)
        base["available_after"] = round(avail - deduct, 4)
        recv_delta = int(round(float(deduct)))
        base["received_before"] = round(recv, 4)
        base["received_after"] = round(float(recv) + float(recv_delta), 4)
        elig, block = _payout_withdraw_eligible(c)
        if not elig:
            base["result"] = "withdraw_blocked"
            prev = base.get("message") or "So du hop le"
            base["message"] = f"{prev}; khong chot duoc: {block}"
            summary["withdraw_blocked"] += 1
        else:
            prev = base.get("message")
            suffix = "; + withdraw_requests CHUA_BAO_KHACH khi chot"
            base["message"] = (prev + suffix) if prev else suffix.strip("; ")
            summary["matched"] += 1
        out_rows.append(base)

    return out_rows, summary


def _mark_duplicates_completed(supabase, duplicate_ids: list[str]) -> int:
    unique = sorted({str(x).strip() for x in duplicate_ids if str(x).strip()})
    if not unique:
        return 0
    updated = 0
    for chunk in _chunked(unique, LOOKUP_CHUNK):
        res = (
            supabase.table(DUPLICATE_TABLE)
            .update({"status": DUPLICATE_STATUS_COMPLETED})
            .in_("id", chunk)
            .eq("status", DUPLICATE_STATUS_PENDING)
            .execute()
        )
        updated += len(res.data or [])
    return updated


@router.get("/withdraw-requests-duplicate")
def list_withdraw_requests_duplicate(
    limit: int = Query(default=500, ge=1, le=2000),
    current_user: dict = Depends(get_current_user),
):
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )
        duplicates = _fetch_pending_duplicates_for_owner(
            supabase, user_owner_global=user_owner, limit=limit
        )
        items = _enrich_duplicate_items(
            supabase, duplicates=duplicates, user_owner_global=user_owner
        )
        min_vnd, _tpl = _fetch_chuyen_tien_settings(supabase)
        filters: dict[str, float | str | None] = {"status": DUPLICATE_STATUS_PENDING}
        if min_vnd > 0:
            filters["min_available_vnd_chuyen_tien"] = round(min_vnd, 4)
        return {"count": len(items), "items": items, "filters": filters}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Query withdraw_requests_duplicate loi: {exc}",
        ) from exc


@router.get("/withdraw-requests-duplicate/export")
def export_withdraw_requests_duplicate_xlsx(
    limit: int = Query(default=2000, ge=1, le=5000),
    current_user: dict = Depends(get_current_user),
):
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )
        duplicates = _fetch_pending_duplicates_for_owner(
            supabase, user_owner_global=user_owner, limit=limit
        )
        items = _enrich_duplicate_items(
            supabase, duplicates=duplicates, user_owner_global=user_owner
        )
        bio = _duplicate_workbook(items)
        data = bio.getvalue()
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="Danh_sach_CK_duplicate.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Export withdraw_requests_duplicate loi: {exc}",
        ) from exc


@router.post("/withdraw-requests-duplicate/payout-file/preview")
async def preview_duplicate_payout_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    try:
        raw = await file.read()
        if len(raw) > PAYOUT_IMPORT_MAX_BYTES:
            raise HTTPException(status_code=400, detail="File qua lon (toi da 5MB)")
        if not raw:
            raise HTTPException(status_code=400, detail="File rong")

        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )

        try:
            df = pd.read_excel(BytesIO(raw), header=0, engine="openpyxl")
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Khong doc duoc Excel: {exc}") from exc

        contacts = _all_zalo_contact_rows_for_owner(supabase, user_owner_global=user_owner)
        by_id_global, _by_id = _contact_maps_for_payout(contacts)
        pending_by_global = _pending_duplicate_map_by_global(
            supabase, user_owner_global=user_owner
        )
        rows, summary = _duplicate_payout_preview_from_dataframe(
            df,
            by_id_global=by_id_global,
            pending_by_global=pending_by_global,
        )

        apply_rows = [
            {
                "contact_id": r["contact_id"],
                "amount_vnd": int(round(float(r["deduct_applied"]))),
                "duplicate_id": r["duplicate_id"],
            }
            for r in rows
            if r.get("result") == "matched"
            and r.get("contact_id") is not None
            and r.get("deduct_applied") is not None
            and float(r["deduct_applied"]) > 0
            and r.get("duplicate_id")
        ]
        loi_apply_rows = [
            {
                "contact_id": r["contact_id"],
                "duplicate_id": r["duplicate_id"],
            }
            for r in rows
            if r.get("result") == "loi_matched"
            and r.get("contact_id") is not None
            and r.get("duplicate_id")
        ]
        return {
            "rows": rows,
            "summary": summary,
            "apply_rows": apply_rows,
            "loi_apply_rows": loi_apply_rows,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Payout preview duplicate loi: {exc}",
        ) from exc


class PayoutApplyItem(BaseModel):
    contact_id: int = Field(..., gt=0)
    amount_vnd: int = Field(..., gt=0)
    duplicate_id: str = Field(..., min_length=1)


class LoiApplyItem(BaseModel):
    contact_id: int = Field(..., gt=0)
    duplicate_id: str = Field(..., min_length=1)


class DuplicatePayoutApplyBody(BaseModel):
    rows: list[PayoutApplyItem] = Field(default_factory=list, max_length=500)
    loi_rows: list[LoiApplyItem] = Field(default_factory=list, max_length=500)


@router.post("/withdraw-requests-duplicate/payout-file/apply")
def apply_duplicate_payout_import(
    body: DuplicatePayoutApplyBody,
    current_user: dict = Depends(get_current_user),
):
    if not body.rows and not body.loi_rows:
        raise HTTPException(status_code=400, detail="Can it nhat mot dong apply hoac loi_rows")
    try:
        supabase = get_supabase_client()
        user_owner = user_owner_global_zalo(current_user)
        if not user_owner:
            raise HTTPException(
                status_code=400,
                detail="Tai khoan khong co id_globalzalo (va khong co id_zl fallback)",
            )

        contacts = _all_zalo_contact_rows_for_owner(supabase, user_owner_global=user_owner)
        _, by_id = _contact_maps_for_payout(contacts)
        pending_by_global = _pending_duplicate_map_by_global(
            supabase, user_owner_global=user_owner
        )
        pending_by_id = {str(d.get("id") or "").strip(): d for d in pending_by_global.values()}

        seen_contacts: set[int] = set()
        duplicate_ids_to_complete: list[str] = []
        applied: list[dict] = []
        loi_applied: list[dict] = []

        for it in body.rows:
            if it.contact_id in seen_contacts:
                raise HTTPException(status_code=400, detail="Trung contact_id trong payload rows")
            seen_contacts.add(it.contact_id)

            dup_id = str(it.duplicate_id).strip()
            dup = pending_by_id.get(dup_id)
            if not dup or str(dup.get("status") or "") != DUPLICATE_STATUS_PENDING:
                raise HTTPException(
                    status_code=400,
                    detail=f"duplicate_id={dup_id} khong ton tai hoac khong con PENDING",
                )

            row = by_id.get(it.contact_id)
            if not row:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} khong ton tai hoac khong thuoc ban",
                )
            id_global = str(row.get("id_global") or "").strip()
            if id_global != str(dup.get("id_global") or "").strip():
                raise HTTPException(
                    status_code=400,
                    detail=f"duplicate_id={dup_id} khong khop id_global contact_id={it.contact_id}",
                )

            amt = int(it.amount_vnd)
            snap = (
                supabase.table("zalo_contacts")
                .select("available_amount,received")
                .eq("id", it.contact_id)
                .limit(1)
                .execute()
            )
            snap_rows = snap.data or []
            if not snap_rows:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} khong doc duoc snapshot",
                )
            cur = snap_rows[0]
            avail = float(cur.get("available_amount") or 0)
            recv = float(cur.get("received") or 0)
            deduct = _payout_deduct_vnd_for_contact(avail=avail, amount_vnd=amt)
            if deduct is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"contact_id={it.contact_id} id_global={id_global} "
                        f"khong du available_amount (con {avail}, can {amt})"
                    ),
                )

            amount_apply = int(round(float(deduct)))
            elig, block = _payout_withdraw_eligible(row)
            if not elig:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"contact_id={it.contact_id} id_global={id_global} "
                        f"khong du dieu kien withdraw_requests: {block}"
                    ),
                )
            bank_fields = _withdraw_bank_fields_from_contact(row)
            assert bank_fields is not None
            bank_type, stk, bank_name = bank_fields

            new_avail = round(avail - float(deduct), 4)
            recv_delta = int(round(float(deduct)))
            new_recv = int(round(float(recv) + float(recv_delta)))

            res = (
                supabase.table("zalo_contacts")
                .update({"available_amount": new_avail, "received": new_recv})
                .eq("id", it.contact_id)
                .execute()
            )
            if not res.data:
                raise HTTPException(
                    status_code=409,
                    detail=f"Khong cap nhat duoc contact_id={it.contact_id}",
                )

            withdraw_id = _insert_withdraw_request_chua_bao_khach(
                supabase,
                id_global=id_global,
                d_name=str(row.get("d_name") or "").strip() or None,
                amount_vnd=amount_apply,
                bank_type=bank_type,
                stk=stk,
                bank_name=bank_name,
            )
            duplicate_ids_to_complete.append(dup_id)
            applied.append(
                {
                    "contact_id": it.contact_id,
                    "id_global": id_global,
                    "duplicate_id": dup_id,
                    "amount_vnd": amt,
                    "deduct_applied": round(float(deduct), 4),
                    "available_after": new_avail,
                    "received_after": new_recv,
                    "withdraw_request_id": withdraw_id,
                    "withdraw_status": WITHDRAW_STATUS_CHUA_BAO_KHACH,
                }
            )

        for it in body.loi_rows:
            dup_id = str(it.duplicate_id).strip()
            dup = pending_by_id.get(dup_id)
            if not dup or str(dup.get("status") or "") != DUPLICATE_STATUS_PENDING:
                raise HTTPException(
                    status_code=400,
                    detail=f"duplicate_id={dup_id} khong ton tai hoac khong con PENDING (loi)",
                )
            row = by_id.get(it.contact_id)
            if not row:
                raise HTTPException(
                    status_code=400,
                    detail=f"contact_id={it.contact_id} khong ton tai hoac khong thuoc ban (loi)",
                )
            id_global = str(row.get("id_global") or "").strip()
            if id_global != str(dup.get("id_global") or "").strip():
                raise HTTPException(
                    status_code=400,
                    detail=f"duplicate_id={dup_id} khong khop id_global (loi)",
                )

            res = (
                supabase.table("zalo_contacts")
                .update({"status_bank": STATUS_BANK_LOI})
                .eq("id", it.contact_id)
                .execute()
            )
            if not res.data:
                raise HTTPException(
                    status_code=409,
                    detail=f"Khong cap nhat LOI_BANK contact_id={it.contact_id}",
                )
            duplicate_ids_to_complete.append(dup_id)
            loi_applied.append(
                {
                    "contact_id": it.contact_id,
                    "id_global": id_global,
                    "duplicate_id": dup_id,
                    "status_bank": STATUS_BANK_LOI,
                }
            )

        completed_count = _mark_duplicates_completed(supabase, duplicate_ids_to_complete)
        return {
            "applied_count": len(applied),
            "loi_applied_count": len(loi_applied),
            "duplicates_completed": completed_count,
            "applied": applied,
            "loi_applied": loi_applied,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Payout apply duplicate loi: {exc}",
        ) from exc
